# -*- coding: utf-8 -*-
"""
BarTender GUI V0.3.3 FIX3k2
- Прогресс-бар на печать всех
- Кнопка "Тестовая страница"
- Строгая валидация дат (ДД.ММ.ГГГГ)
- ShortName берётся только из Excel (столбец ShortName / SHORTNAME / Короткое имя)
"""

import os, csv, re, time, json, atexit, traceback, datetime as dt
import time
import customtkinter as ctk
from tkinter import filedialog as fd, messagebox as mb, Menu
from PIL import Image
import sys
# --- Paths relative to app root (works for .py and PyInstaller .exe) ---
def _app_base_dir():
    try:
        if getattr(sys, "frozen", False):
            return os.path.dirname(sys.executable)   # folder of .exe
        return os.path.dirname(os.path.abspath(__file__))  # folder of .py
    except Exception:
        return os.getcwd()

BASE_DIR = _app_base_dir()
PRODUCT_MAP_DEFAULT = os.path.join(BASE_DIR, "Список товаров.xlsx")

try:
    from win32com.client import Dispatch
    import win32print
except Exception:
    Dispatch = None
    win32print = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

from calendar import monthrange

APP_TITLE    = "BarTender GUI V2.0 batch"
APP_VERSION = "2.0"
PREVIEW_NAME = "preview.png"
AUTO_MARKING_ENABLED = True
MARKING_PRINTER_NAME = ""
MARKING_LABEL_TEMPLATE_PATH = os.path.join(BASE_DIR, "templates", "marking_label.btw")

# ------------------------ Конфиг ------------------------

def _cfg_dir():
    base = os.path.join(os.environ.get("APPDATA", os.getcwd()), "BarTenderGUI")
    os.makedirs(base, exist_ok=True)
    return base

def _cfg_path(): 
    return os.path.join(_cfg_dir(), "config.json")

def load_config():
    p = _cfg_path()
    if os.path.isfile(p):
        try:
            with open(p, "r", encoding="utf-8") as f: 
                return json.load(f)
        except Exception: 
            return {}
    return {}

def save_config(cfg: dict):
    try:
        with open(_cfg_path(), "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

# ------------------------ Утилиты ------------------------

def enum_printers():
    if not win32print: 
        return []
    flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
    try:
        return [name for (_f,_d,name,_c) in win32print.EnumPrinters(flags)]
    except Exception:
        return []

def only_digits(s): 
    return re.sub(r"\D+","", s or "")

def short_gtin(gtin): 
    d=only_digits(gtin); 
    return d[-3:].zfill(3) if d else ""

def parse_date_ru(s: str):
    """Строгий парсер ДД.ММ.ГГГГ → date, иначе ValueError."""
    s = (s or "").strip()
    if not s:
        raise ValueError("Пустая дата")
    try:
        return dt.datetime.strptime(s, "%d.%m.%Y").date()
    except Exception:
        raise ValueError(f"Неверный формат даты: '{s}', нужен ДД.ММ.ГГГГ")

def parse_prod_date(s: str):
    """Мягкий парсер (используется при чтении сохранённого): вернёт date или None."""
    s=(s or "").strip()
    for fmt in ("%d.%m.%Y","%Y-%m-%d"):
        try: 
            return dt.datetime.strptime(s, fmt).date()
        except Exception: 
            pass
    return None

def add_days(date_, days:int): 
    return date_ + dt.timedelta(days=days)

def add_months(date_, months:int):
    y = date_.year + (date_.month - 1 + months) // 12
    m = (date_.month - 1 + months) % 12 + 1
    d = min(date_.day, monthrange(y, m)[1])
    return dt.date(y, m, d)

def add_years(date_, years:int):
    y = date_.year + years
    d = min(date_.day, monthrange(y, date_.month)[1])
    return dt.date(y, date_.month, d)

# ------------------------ CSV (Контур сырой) ------------------------

def load_kontur_raw(csv_path: str):
    rows = []
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        rdr = csv.reader(f, delimiter="\t", quotechar='"')
        for parts in rdr:
            if not parts or all((p or "").strip()=="" for p in parts): 
                continue
            parts = [(p or "").strip() for p in parts]
            while len(parts)<3: 
                parts.append("")
            dm, gtin, name = parts[0], parts[1], parts[2]
            if dm.strip().upper()=="DM" and (name.strip().upper() in ("NAME","")):
                continue
            if not dm.strip(): 
                continue
            rows.append({"DM": dm, "GTIN": gtin, "NAME": name})
    return rows

# ------------------------ Excel-справочник ------------------------

def _norm(s): 
    return re.sub(r"\s+","", (s or "").strip().lower())

def _parse_shelf_life(text: str):
    """{'days'| 'months' | 'years' | 'weeks' | 'raw'}"""
    t = (text or "").strip().lower()
    res = {'days':None, 'months':None, 'years':None, 'weeks':None, 'raw':text}
    if not t: 
        return res
    m = re.search(r"(\d+)", t)
    if not m: 
        return res
    n = int(m.group(1))
    if any(w in t for w in ["год","лет","года"]):
        res['years'] = n; return res
    if any(w in t for w in ["мес","месяц","месяцев","месяца"]):
        res['months'] = n; return res
    if any(w in t for w in ["нед","неделя","недели","недель"]):
        res['weeks'] = n; return res
    if any(w in t for w in ["сут","дн"]):
        res['days'] = n; return res
    res['months'] = n; 
    return res

def read_product_map(xlsx_path: str):
    """Читает: GTIN, Упаковка, Срок годности, Шаблон партии, ShortName."""
    mapping={}
    if not xlsx_path or not os.path.isfile(xlsx_path) or not load_workbook: 
        return mapping
    try:
        wb=load_workbook(xlsx_path, data_only=True); ws=wb.active
        hdrs=[str(c.value or "").strip() for c in ws[1]]
        col={h:i+1 for i,h in enumerate(hdrs)}

        idx_gtin = col.get("GTIN") or col.get("ГТИН") or col.get("gtin")
        idx_pack = col.get("Упаковка") or col.get("УПАКОВКА")
        idx_shelf= col.get("Срок годности") or col.get("СРОК ГОДНОСТИ")
        idx_part = col.get("Шаблон партии") or col.get("ШАБЛОН ПАРТИИ") or col.get("Партия") or col.get("ПАРТИЯ")
        idx_short = (col.get("ShortName") or col.get("SHORTNAME") or col.get("Shortname") or col.get("shortname")
                     or col.get("Short Name") or col.get("SHORT NAME") or col.get("Короткое имя") or col.get("КОРОТКОЕ ИМЯ"))

        def cell(row, j):
            if not j: return ""
            v=row[j-1].value
            return "" if v is None else str(v).strip()

        has_short = bool(idx_short)
        for row in ws.iter_rows(min_row=2):
            g=only_digits(cell(row, idx_gtin))
            if not g: 
                continue
            pack=_norm(cell(row, idx_pack))
            shelf=cell(row, idx_shelf)
            fmt=""
            if "ведро" in pack: 
                fmt="30x20"
            elif any(w in pack for w in ("банка","туба")): 
                fmt="16x16"
            shelf_parsed=_parse_shelf_life(shelf)
            part_tpl=cell(row, idx_part)
            short_from_xlsx = cell(row, idx_short)
            mapping[g]={"FORMAT":fmt, "SHELF":shelf_parsed, "PART_TEMPLATE":part_tpl, "SHORTNAME": short_from_xlsx}
        mapping["_HAS_SHORT_COL"] = has_short
        return mapping
    except Exception:
        return {}

def choose_format_for(gtin, product_map, manual_choice):
    if manual_choice in ("16x16","30x20"): 
        return manual_choice
    info = product_map.get(only_digits(gtin), {})
    fmt = (info.get("FORMAT") or "") if info else ""
    return fmt if fmt in ("16x16","30x20") else "16x16"

def make_part_num(prod_date, part_template: str|None):
    yymmdd = prod_date.strftime("%y%m%d") if prod_date else ""
    if part_template: 
        return part_template.replace("{DATE}", yymmdd)
    return yymmdd

def enrich_row(base_row, idx1, prod_date, exp_days_override, product_map, mode_choice):
    dm,gtin,name = base_row.get("DM",""), base_row.get("GTIN",""), base_row.get("NAME","")
    fmt = choose_format_for(gtin, product_map, mode_choice)
    info = product_map.get(only_digits(gtin), {})
    pd = prod_date or dt.date.today()

    # срок годности
    exp_date = None
    shelf_info = info.get("SHELF") or {}
    log_shelf = ""
    if shelf_info:
        if shelf_info.get('years'):
            exp_date = add_years(pd, int(shelf_info['years'])); log_shelf = f"years={shelf_info['years']}"
        elif shelf_info.get('months'):
            exp_date = add_months(pd, int(shelf_info['months'])); log_shelf = f"months={shelf_info['months']}"
        elif shelf_info.get('weeks'):
            exp_date = add_days(pd, int(shelf_info['weeks'])*7); log_shelf = f"weeks={shelf_info['weeks']}"
        elif shelf_info.get('days'):
            exp_date = add_days(pd, int(shelf_info['days'])); log_shelf = f"days={shelf_info['days']}"
    if (exp_date is None) and (exp_days_override is not None):
        exp_date = add_days(pd, int(exp_days_override)); log_shelf = f"override_days={exp_days_override}"

    part_num = make_part_num(pd, info.get("PART_TEMPLATE") or "")
    short_excel = (info.get("SHORTNAME") or "").strip()

    enr = {
        "DM": dm, "GTIN": gtin, "NAME": name,
        "ShortGTIN": short_gtin(gtin),
        "ShortName": short_excel if short_excel else name,
        "PROD_DATE": pd.strftime("%d.%m.%Y") if pd else "",
        "NUM": str(idx1),
        "EXP_DATE": exp_date.strftime("%d.%m.%Y") if exp_date else "",
        "PART_NUM": part_num,
        "_FORMAT": fmt,
        "_SHELF_LOG": log_shelf,
        "_SHORT_SRC": "Excel" if short_excel else "NAME"
    }
    return enr

# ------------------------ COM-обёртка ------------------------


# === PATCH HELPERS (batch printing) ===
def _dbg(s): 
    try:
        self.logger.log(s)
    except Exception:
        print(s)

def _get_pack_size(self):
    """Возвращает размер пакета из GUI (поле рядом с 'Пакет').
    Пишем диагностику значения и fallback к 0."""
    gui_val_raw = None
    gui_val_int = 0
    try:
        if hasattr(self, "batch_entry") and self.batch_entry:
            gui_val_raw = self.batch_entry.get()
            gui_val_int = int(str(gui_val_raw).strip() or "0")
    except Exception as e:
        _dbg(f"[DEBUG] PACK GUI parse failed: raw={gui_val_raw!r} err={e}")
        gui_val_int = 0
    try:
        _dbg(f"[DEBUG] PACK from GUI: raw={gui_val_raw!r} -> int={gui_val_int}")
    except Exception:
        pass
    return max(0, gui_val_int)
# === END PATCH HELPERS ===
class BT:
    def __init__(self, logger): 
        self.logger=logger
        self.app=None

    def start(self):
        if not Dispatch: 
            raise RuntimeError("pywin32 не установлен")
        self.logger.log("Запуск BarTender COM...")
        self.app=Dispatch("BarTender.Application"); 
        self.app.Visible=False
        self.logger.log("BarTender COM запущен.")

    def stop(self):
        if self.app:
            try: 
                self.logger.log("Завершение BarTender COM...")
                self.app.Quit(1)
            except Exception: 
                pass
            self.app=None

    def open_format(self, path):
        self.logger.log(f"Открытие шаблона: {path}")
        fmt=self.app.Formats.Open(path, False, "")
        try: 
            self.logger.log(f"NamedSubStrings: {[s.Name for s in fmt.NamedSubStrings]}")
        except Exception: 
            pass
        return fmt

    def set_common_print_flags(self, fmt):
        for a,v in (("UseDatabase",False),("SelectRecordsAtPrint",False),("RecordRange","1")):
            try: setattr(fmt,a,v)
            except Exception: pass

    def apply_fields(self, fmt, data:dict):
        names=set()
        try: names={s.Name for s in fmt.NamedSubStrings}
        except Exception: pass
        payload={k:v for k,v in data.items() if not k.startswith("_") and ((not names) or (k in names))}
        skipped=sorted(set(data.keys())-set(payload.keys())-{k for k in data if k.startswith("_")})
        if skipped: 
            self.logger.log(f"Подстановка: пропущены поля (нет в шаблоне): {skipped}")
        cnt=0
        for k,v in payload.items():
            try: 
                fmt.SetNamedSubStringValue(k, str(v)); 
                cnt+=1; 
                continue
            except Exception: 
                pass
            try:
                subs=getattr(fmt,"SubStrings",None)
                if subs: 
                    subs(k).Value=str(v); 
                    cnt+=1
            except Exception: 
                pass
        self.logger.log(f"Подстановка полей: всего={len(payload)}, успешно={cnt}")
        return cnt>0

    def export_preview(self, fmt, path):
        try: 
            fmt.ExportToFile(path, "PNG", 1, 300, 0)
            return True
        except Exception: 
            return False

# ------------------------ Логгер ------------------------


class Logger:
    def __init__(self, tb):
        self.tb = tb
        try:
            self.tb.tag_configure("info", foreground="orange")
            self.tb.tag_configure("pack", foreground="blue")
            self.tb.tag_configure("error", foreground="red")
        except Exception:
            pass

    def log(self, msg):
        ts = time.strftime("%H:%M:%S")
        tag = None
        if "[INFO]" in msg:
            tag = "info"
        elif "[PACK]" in msg:
            tag = "pack"
        elif "ERROR" in msg:
            tag = "error"
        self.tb.configure(state="normal")
        try:
            self.tb.insert("end", f"[{ts}] {msg}\n", tag)
        except Exception:
            self.tb.insert("end", f"[{ts}] {msg}\n")
        self.tb.see("end")
        self.tb.configure(state="normal")
        self.tb.update_idletasks()

    def err(self, msg):
        self.log(f"ERROR: {msg}")

class App(ctk.CTk):


    # ---------- helper: normalized checkbox value ----------
    def _dialog_flag(self) -> bool:
        """Читает чекбокс «Показывать диалог BarTender» как bool надёжно."""
        try:
            v = self.show_dialog_var.get()
        except Exception:
            try:
                v = self.default_show_dialog
            except Exception:
                return False
        try:
            if isinstance(v, str):
                s = v.strip().lower()
                if s.isdigit():
                    return bool(int(s))
                return s in ("true","yes","on","1")
            return bool(v)
        except Exception:
            return False

    def _write_tmp_batch_csv(self, rows_enriched, path=None):
        import os, csv
        cols = getattr(self, "REQ_COLS", ["ShortName","ShortGTIN","EXP_DATE","PROD_DATE","PART_NUM","DM","NUM"])
        if path is None:
            path = os.path.join(BASE_DIR, "tmp_batch.csv")
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
        except Exception:
            pass
        # Сформируем список строк: при калибровке добавим 6 тестовых в начало
        rows2 = list(rows_enriched)
        try:
            if bool(self.calib_var.get()):
                dummy = {k: ("1" if k.upper()=="NUM" else ("000" if k=="ShortGTIN" else "X")) for k in cols}
                rows2 = [dummy.copy() for _ in range(6)] + rows2
        except Exception:
            pass
        try:
            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                w = csv.DictWriter(f, fieldnames=cols, delimiter=";")
                w.writeheader()
                for enr in rows2:
                    row = {k: (enr.get(k, "") or "") for k in cols}
                    w.writerow(row)
            if getattr(self, "logger", None):
                try:
                    self.logger.log(f"tmp_batch.csv записан: {path} (строк={len(rows2)})")
                except Exception:
                    pass
        except Exception as e:
            try:
                if getattr(self, "logger", None):
                    self.logger.err(f"Не удалось записать tmp_batch.csv: {e}")
            except Exception:
                pass
        return path


    def _rangecsv_repoint_db(self, *args, **kwargs):
        try: self.logger.log("[TMPBATCH] DB rebind отключен (stub)")
        except Exception: pass
        return False

    def __init__(self):
        self.cancel_requested = False
        self.is_paused = False
        super().__init__()
        self.REQ_COLS = ["ShortName","ShortGTIN","EXP_DATE","PROD_DATE","PART_NUM","DM","NUM"]
        self.title(APP_TITLE)
        self.geometry("1300x900")
        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("blue")

        self.cfg = load_config()
        # Параметры по умолчанию
        self.default_batch_size = int(self.cfg.get('batch_size', 1830))
        self.default_show_dialog = bool(self.cfg.get('show_print_dialog', False))
        atexit.register(lambda: save_config(self.cfg))

        self.bt = None
        self.is_paused = False
        self.csv_path = ""
        self.csv_rows = []
        self.preview_ctkimg = None
        self.product_map = {}

        self._build_ui()
        self._start_bt()
        self._refresh_printers()
        self._load_presets()
        self._auto_load_product_map()
        self.after(200, self._prompt_csv_on_launch)

    # индикатор прогресса
    
    def _set_progress(self, cur: int, total: int, phase: str = ""):
        try:
            if total <= 0:
                self.progress_bar.set(0.0)
                self.progress_label.configure(text=phase or "Готово")
            else:
                frac = max(0.0, min(1.0, cur / total))
                self.progress_bar.set(frac)
                self.progress_label.configure(text=f"{phase} {cur}/{total}")
            self.update_idletasks()
            try:
                self.update()
            except Exception:
                pass
        except Exception:
            pass

    def _build_ui(self):
        # верхняя панель: выбор формата/шаблонов/CSV
        top = ctk.CTkFrame(self, corner_radius=12)
        top.pack(fill="x", padx=12, pady=(12, 6))

        ctk.CTkLabel(top, text="Формат (Auto/ручной):").pack(side="left", padx=(8, 4), pady=10)
        self.format_combo = ctk.CTkComboBox(top, values=["Auto", "16x16", "30x20"], state="readonly", width=120)
        self.format_combo.set("Auto")
        self.format_combo.pack(side="left", padx=(0, 10), pady=10)
        ctk.CTkLabel(top, text="Шаблоны: авто (корень программы)").pack(side="left", padx=(4, 8), pady=10)

        self.btw16_entry = ctk.CTkEntry(top, placeholder_text="BTW для 16x16")
#         self.btw16_entry.pack(side="left", fill="x", expand=True, padx=(4, 4), pady=10)
#         ctk.CTkButton(top, text="16x16…", command=lambda: self._choose_btw("16x16")).pack(side="left", padx=(0, 8), pady=10)

        self.btw30_entry = ctk.CTkEntry(top, placeholder_text="BTW для 30x20")
#         self.btw30_entry.pack(side="left", fill="x", expand=True, padx=(4, 4), pady=10)
#         ctk.CTkButton(top, text="30x20…", command=lambda: self._choose_btw("30x20")).pack(side="left", padx=(0, 8), pady=10)

        ctk.CTkButton(top, text="Открыть CSV…", command=self._choose_csv).pack(side="left", padx=(8, 8), pady=10)
        self.csv_label = ctk.CTkLabel(top, text="CSV: (не выбран)")
        self.csv_label.pack(side="left", padx=(8, 4), pady=10)

        # справочник
        prod = ctk.CTkFrame(self, corner_radius=12)
#         prod.pack(fill="x", padx=12, pady=6)
        ctk.CTkLabel(prod, text="Список товаров.xlsx (автозагрузка из C:\\auto_print):").pack(side="left", padx=(8, 6), pady=10)
        self.prodmap_entry = ctk.CTkEntry(prod, placeholder_text=PRODUCT_MAP_DEFAULT)
        self.prodmap_entry.pack(side="left", fill="x", expand=True, padx=(4, 4), pady=10)
        self.prodmap_entry.configure(state="disabled")

        # параметры печати
        mid = ctk.CTkFrame(self, corner_radius=12)
        mid.pack(fill="x", padx=12, pady=6)

        ctk.CTkLabel(mid, text="Принтер:").pack(side="left", padx=(12, 6), pady=10)
        self.prn_combo = ctk.CTkComboBox(mid, values=["(нет принтеров)"], state="readonly", width=320)
        self.prn_combo.pack(side="left", padx=(0, 12), pady=10)

        ctk.CTkLabel(mid, text="Строка № (1-based):").pack(side="left", padx=(6, 6), pady=10)
        self.index_entry = ctk.CTkEntry(mid, width=80)
        self.index_entry.insert(0, "1")
        self.index_entry.pack(side="left", padx=(0, 12), pady=10)

        ctk.CTkLabel(mid, text="Копий/шт:").pack(side="left", padx=(6, 6), pady=10)
        self.copies_entry = ctk.CTkEntry(mid, width=80)
        self.copies_entry.insert(0, "1")
        self.copies_entry.pack(side="left", padx=(0, 12), pady=10)

        # Параметры печати: лимит и размер батча
        ctk.CTkLabel(mid, text="Лимит (печать всех):").pack(side="left", padx=(6, 6), pady=10)
        self.limit_entry = ctk.CTkEntry(mid, width=100)
        self.limit_entry.pack(side="left", padx=(0, 12), pady=10)

        ctk.CTkLabel(mid, text="Пакет по (шт.):").pack(side="left", padx=(6, 6), pady=10)
        self.batch_entry = ctk.CTkEntry(mid, width=100)
        self.batch_entry.insert(0, str(self.default_batch_size))
        self.batch_entry.pack(side="left", padx=(0, 12), pady=10)

        self.show_dialog_var = ctk.BooleanVar(value=self.default_show_dialog)
        self.show_dialog_chk = ctk.CTkCheckBox(mid, text="Показывать диалог печати BarTender", variable=self.show_dialog_var)
        self.show_dialog_chk.pack(side="left", padx=(6, 12), pady=10)
        self.show_dialog_chk.pack(side="left", padx=(6, 12), pady=10)

        self.calib_var = ctk.BooleanVar(value=False)
        self.calib_chk = ctk.CTkCheckBox(mid, text="Калибровка TSC: 6 этикеток 'X' перед печатью", variable=self.calib_var)
        self.calib_chk.pack(side="left", padx=(12, 12), pady=10)

        # блок обогащения: даты + партия
        enr = ctk.CTkFrame(self, corner_radius=12)
        enr.pack(fill="x", padx=12, pady=6)

        ctk.CTkLabel(enr, text="Дата произв. (ДД.ММ.ГГГГ):").pack(side="left", padx=(12, 6), pady=10)
        self.prod_date_entry = ctk.CTkEntry(enr, width=140)
        self.prod_date_entry.insert(0, dt.date.today().strftime("%d.%m.%Y"))
        self.prod_date_entry.pack(side="left", padx=(0, 6), pady=10)
        ctk.CTkButton(enr, text="📅", width=36, command=self._open_date_picker).pack(side="left", padx=(0, 12), pady=10)
        vcmd = (self.register(lambda s: re.fullmatch(r"[0-9.]*", s or "") is not None), "%P")
        self.prod_date_entry.configure(validate="key", validatecommand=vcmd)

#         ctk.CTkLabel(enr, text="Срок годности, дней (если не в Excel):").pack(side="left", padx=(6, 6), pady=10)
        self.exp_days_entry = ctk.CTkEntry(enr, width=120)
#         self.exp_days_entry.pack(side="left", padx=(0, 12), pady=10)

        self.part_auto_var = ctk.BooleanVar(value=False)
        self.part_auto_chk = ctk.CTkCheckBox(enr, text="Авто-партия (YYMMDD)", variable=self.part_auto_var)
        self.part_auto_chk.pack(side="left", padx=(12, 6), pady=10)

        self.part_entry = ctk.CTkEntry(enr, width=160, placeholder_text="Номер партии вручную")
        self.part_entry.pack(side="left", padx=(0, 12), pady=10)

        # кнопки действий
        btns = ctk.CTkFrame(self)
        btns.pack(fill="x", padx=8, pady=(4, 8))
        ctk.CTkButton(btns, text="Превью выбранной строки", command=self._preview, height=36).pack(side="left", padx=6, pady=8)
        ctk.CTkButton(btns, text="Печать первой строки", command=self._print_one, height=36).pack(side="left", padx=6, pady=8)
        ctk.CTkButton(btns, text="Печать пакетами", command=self._print_range_one_job_via_csv, height=36).pack(side="left", padx=6, pady=8)

        ctk.CTkButton(btns, text="Тестовая страница", command=self._print_test, height=36).pack(side="left", padx=6, pady=8)

        self.pause_btn = ctk.CTkButton(btns, text="Пауза", command=self._toggle_pause, height=36)
        self.pause_btn.pack(side="left", padx=6, pady=8)
        # Кнопка отмены печати
        self.cancel_btn = ctk.CTkButton(btns, text="Отмена печати", command=self._cancel_print, height=36)
        self.cancel_btn.pack(side="left", padx=6, pady=8)

# прогресс
        pframe = ctk.CTkFrame(self, corner_radius=8)
        pframe.pack(fill="x", padx=12, pady=(0, 8))
        self.progress_bar = ctk.CTkProgressBar(pframe)
        self.progress_bar.pack(fill="x", padx=12, pady=(8, 4))
        self.progress_bar.set(0.0)
        self.progress_label = ctk.CTkLabel(pframe, text="Готово")
        self.progress_label.pack(anchor="w", padx=12, pady=(0, 8))

        # низ: превью + лог
        bottom = ctk.CTkFrame(self, corner_radius=12)
        bottom.pack(fill="both", expand=True, padx=12, pady=(6, 12))

        self.preview_label = ctk.CTkLabel(bottom, text="Предпросмотр будет здесь", width=560, height=360, corner_radius=8)
        self.preview_label.pack(side="left", fill="both", expand=True, padx=12, pady=12)

        right = ctk.CTkFrame(bottom, corner_radius=12)
        right.pack(side="left", fill="both", expand=True, padx=(0, 12), pady=12)

        self.logbox = ctk.CTkTextbox(right, height=360)
        self.logbox.pack(fill="both", expand=True, padx=12, pady=12)
        self.logger = Logger(self.logbox)
        self._enable_log_copy()

    def _enable_log_copy(self):
        def _copy():
            try:
                sel = self.logbox.get("sel.first", "sel.last")
            except Exception:
                sel = ""
            if sel:
                self.clipboard_clear()
                self.clipboard_append(sel)
            return "break"

        def _all():
            try:
                self.logbox.tag_add("sel", "1.0", "end-1c")
            except Exception:
                pass
            return "break"

        self.logbox.bind("<Control-c>", lambda e: _copy())
        self.logbox.bind("<Control-a>", lambda e: _all())

        m = Menu(self, tearoff=0)
        m.add_command(label="Копировать", command=_copy)
        m.add_command(label="Выделить всё", command=_all)
        m.add_separator()
        m.add_command(label="Сохранить лог…", command=self._save_log)
        self.logbox.bind("<Button-3>", lambda e: (m.tk_popup(e.x_root, e.y_root), m.grab_release()))

    def _start_bt(self):
        try:
            self.bt = BT(self.logger)
            self.bt.start()
        except Exception as e:
            self.logger.err(f"COM не стартовал: {e}")
            mb.showerror("BarTender COM", f"Не удалось запустить COM:\n{e}")

    def _refresh_printers(self):
        items = enum_printers() or ["(нет принтеров)"]
        self.prn_combo.configure(values=items)
        self.prn_combo.set(items[0])

    def _load_presets(self):
        p16 = (self.cfg.get("formats", {}) or {}).get("16x16", "")
        p30 = (self.cfg.get("formats", {}) or {}).get("30x20", "")
        if p16 and os.path.isfile(p16):
            self.btw16_entry.insert(0, p16)
        if p30 and os.path.isfile(p30):
            self.btw30_entry.insert(0, p30)

    def _auto_load_product_map(self):

        path = PRODUCT_MAP_DEFAULT
        # если нет файла в корне — попросим указать и скопируем в корень под стандартным именем
        if not os.path.isfile(path):
            try:
                self.logger.err(f"Справочник не найден: {path}. Укажи 'Список товаров.xlsx' (будет скопирован в корень).")
            except Exception:
                pass
            p = fd.askopenfilename(title="Укажи Список товаров.xlsx", filetypes=[("Excel", "*.xlsx")])
            if p and os.path.isfile(p):
                try:
                    os.makedirs(BASE_DIR, exist_ok=True)
                except Exception:
                    pass
                try:
                    import shutil as _shutil
                    _shutil.copy2(p, path)
                except Exception:
                    path = p  # если копия не удалась — используем выбранный путь напрямую

        try:
            self.prodmap_entry.configure(state="normal")
            self.prodmap_entry.delete(0, "end")
            self.prodmap_entry.insert(0, path)
            self.prodmap_entry.configure(state="disabled")
        except Exception:
            pass

        if os.path.isfile(path) and load_workbook is not None:
            try:
                self.product_map = read_product_map(path) or {}
            except Exception as e:
                self.product_map = {}
                self.logger.err(f"Ошибка чтения справочника: {e}")
            self.cfg["product_map_path"] = path
            save_config(self.cfg)
            cnt = len([k for k in self.product_map.keys() if k != "_HAS_SHORT_COL"])
            has_short = bool(self.product_map.get("_HAS_SHORT_COL"))
            self.logger.log(f"Справочник загружен автоматически: {path} (записей={cnt}); ShortName-столбец: {has_short}")
        else:
            if load_workbook is None:
                self.logger.err("openpyxl не установлен; справочник не загружен.")
            else:
                self.logger.err(f"Справочник не найден: {path}")

    def _prompt_csv_on_launch(self):
        if self.csv_path:
            return
        p = fd.askopenfilename(title="Выбери kontur.csv/tsv",
                               filetypes=[("CSV/TSV", "*.csv;*.tsv;*.txt"), ("Все файлы", "*.*")])
        if not p:
            self.logger.err("CSV не выбран.")
            return
        self.csv_path = p
        self.csv_label.configure(text=f"CSV: {self.csv_path}")
        try:
            self.csv_rows = load_kontur_raw(p)
            head = self.csv_rows[0] if self.csv_rows else {"DM": "", "NAME": ""}
            self.logger.log(f"CSV: строк={len(self.csv_rows)}; пример DM='{head.get('DM','')}', NAME='{head.get('NAME','')}'")
            # авто-подстановка размера пакета: минимум из дефолта и общего числа строк
            try:
                if self.csv_rows:
                    suggested = min(len(self.csv_rows), self.default_batch_size)
                    self.batch_entry.delete(0, "end")
                    self.batch_entry.insert(0, str(suggested))
            except Exception:
                pass

            # ---- авто-превью первой строки сразу после выбора CSV (старт программы) ----
            try:
                if self.csv_rows:
                    try:
                        self.index_entry.delete(0, 'end')
                        self.index_entry.insert(0, '1')
                    except Exception:
                        pass
                    self._preview()
            except Exception as _e:
                try:
                    self.logger.err(f"Авто-превью при запуске: {_e}")
                except Exception:
                    pass
            # ---- конец авто-превью ----
        except Exception as e:
            self.logger.err(f"Ошибка CSV: {e}")
            mb.showerror("CSV", f"Не удалось прочитать файл:\n{e}")

    def _choose_btw(self, fmt):
        p = fd.askopenfilename(title=f"Выбери BTW для {fmt}", filetypes=[("BarTender Template", "*.btw")])
        if not p:
            return
        if fmt == "16x16":
            self.btw16_entry.delete(0, "end")
            self.btw16_entry.insert(0, p)
        else:
            self.btw30_entry.delete(0, "end")
            self.btw30_entry.insert(0, p)
        self.cfg.setdefault("formats", {})[fmt] = p
        save_config(self.cfg)
        self.logger.log(f"[Пресет] Сохранён путь для {fmt}: {p}")

    def _choose_csv(self):
        p = fd.askopenfilename(title="Выбери kontur.csv/tsv",
                               filetypes=[("CSV/TSV", "*.csv;*.tsv;*.txt"), ("Все файлы", "*.*")])
        if not p:
            return
        self.csv_path = p
        self.csv_label.configure(text=f"CSV: {self.csv_path}")
        try:
            self.csv_rows = load_kontur_raw(p)
            head = self.csv_rows[0] if self.csv_rows else {"DM": "", "NAME": ""}
            self.logger.log(f"CSV: строк={len(self.csv_rows)}; пример DM='{head.get('DM','')}', NAME='{head.get('NAME','')}'")
            # авто-подстановка размера пакета: минимум из дефолта и общего числа строк
            try:
                if self.csv_rows:
                    suggested = min(len(self.csv_rows), self.default_batch_size)
                    self.batch_entry.delete(0, "end")
                    self.batch_entry.insert(0, str(suggested))
            except Exception:
                pass

            # ---- авто-превью первой строки сразу после выбора CSV ----
            try:
                if self.csv_rows:
                    # сбросим индекс на 1 и вызовем превью
                    try:
                        self.index_entry.delete(0, 'end')
                        self.index_entry.insert(0, '1')
                    except Exception:
                        pass
                    self._preview()
            except Exception as _e:
                # не падать из-за превью — просто залогируем
                try:
                    self.logger.err(f"Авто-превью после выбора CSV: {_e}")
                except Exception:
                    pass
            # ---- конец авто-превью ----
        except Exception as e:
            self.logger.err(f"Ошибка CSV: {e}")
    # ---------- helpers ----------
    def _get_batch_size(self):
        t = (self.batch_entry.get() or "").strip()
        if not t:
            return None
        try:
            v = max(1, int(t))
            return v
        except Exception:
            mb.showerror("Пакет", f"Ожидается целое число, получено: '{t}'")
            return None


    def _get_printer(self):
        prn = self.prn_combo.get()
        if not prn or prn == "(нет принтеров)":
            mb.showerror("Принтер", "Принтер не выбран.")
            return None
        return prn

    def _get_prod_date(self):
        try:
            txt = (self.prod_date_entry.get() or "").strip()
        except Exception:
            txt = ""

        try:
            return parse_date_ru(txt)
        except Exception as e:
            mb.showerror("Дата производства", str(e))
            return None

    def _get_exp_days(self):
        t = (self.exp_days_entry.get() or "").strip()
        if not t:
            return None
        try:
            return max(0, int(t))
        except Exception:
            mb.showerror("Срок годности (дней)", f"Ожидается целое число дней, получено: '{t}'")
            return None

    def _get_index(self):
        t = (self.index_entry.get() or "").strip()
        try:
            i = int(t or "1")
        except Exception:
            i = 1
        return max(1, i)

    def _get_limit(self):
        t = (self.limit_entry.get() or "").strip()
        if t == "" or t == "0":
            return None
        try:
            v = int(t)
            if v <= 0:
                return None
            return v
        except Exception:
            mb.showerror("Лимит", f"Ожидается целое число, получено: '{t}'")
            return None
        try:
            return max(1, int(t))
        except Exception:
            mb.showerror("Лимит", f"Ожидается целое число, получено: '{t}'")
            return None

    def _get_copies(self):
        t = (self.copies_entry.get() or "").strip()
        try:
            return max(1, int(t or "1"))
        except Exception:
            mb.showerror("Копий/шт", f"Ожидается целое число, получено: '{t}'")
            return 1


    def _get_btw_for_format(self, fmt_name: str):

        """

        BTW ищем рядом с .exe/.py (BASE_DIR). Если нет — спрашиваем путь и копируем выбранный файл

        в корень программы под именем по умолчанию (16x16.btw/30x20.btw), чтобы exe всё "запомнил".

        """

        default_name = "16x16.btw" if fmt_name == "16x16" else "30x20.btw"

        default_path = os.path.join(BASE_DIR, default_name)

        if os.path.isfile(default_path):

            return default_path

    

        # если ранее в полях что-то было — используем

        try:

            path = (self.btw16_entry.get().strip() if fmt_name == "16x16" else self.btw30_entry.get().strip())

        except Exception:

            path = ""

        if path and os.path.isfile(path):

            return path

    

        try:

            if getattr(self, "logger", None):

                self.logger.err(f"BTW не найден в корне: {default_path}. Укажи путь к шаблону {fmt_name}.")

        except Exception:

            pass

    

        p = fd.askopenfilename(title=f"Укажи BTW для {fmt_name}", filetypes=[("BarTender Template", "*.btw")])

        if p and os.path.isfile(p):

            # копируем в корень программы под стандартным именем

            try:

                os.makedirs(BASE_DIR, exist_ok=True)

            except Exception:

                pass

            try:

                target = os.path.join(BASE_DIR, default_name)

                import shutil as _shutil

                _shutil.copy2(p, target)

                # запишем в поля на всякий случай

                try:

                    if fmt_name == "16x16":

                        self.btw16_entry.delete(0, "end"); self.btw16_entry.insert(0, target)

                    else:

                        self.btw30_entry.delete(0, "end"); self.btw30_entry.insert(0, target)

                except Exception:

                    pass

                return target

            except Exception:

                return p

    

        mb.showerror("BTW", f"Шаблон для формата {fmt_name} не найден.")

        return None

    
    def _enrich(self, base_row, idx1):
        d = self._get_prod_date()
        if not d:
            return None
        enr = enrich_row(
            base_row=base_row,
            idx1=idx1,
            prod_date=d,
            exp_days_override=self._get_exp_days(),
            product_map=self.product_map,
            mode_choice=self.format_combo.get()
        )

        # ручная партия имеет приоритет (если авто-галка снята)
        if not self.part_auto_var.get():
            manual = (self.part_entry.get() or "").strip()
            enr["PART_NUM"] = manual

        gtin_key = only_digits(base_row.get("GTIN", ""))
        info = self.product_map.get(gtin_key, {})
        shelf = info.get("SHELF") or {}
        shelf_desc = shelf.get("raw") or "-"
        self.logger.log(
            f"GTIN lookup: FORMAT='{info.get('FORMAT','-') or '-'}', "
            f"SHELF='{shelf_desc}', SHELF_PARSED={enr.get('_SHELF_LOG','')}, "
            f"PART_TPL={'-' if not info.get('PART_TEMPLATE') else info.get('PART_TEMPLATE')}, "
            f"ShortNameExcel='{(info.get('SHORTNAME') or '').strip()}', GTIN={gtin_key}"
        )

        src = "не задан"
        if enr.get("_SHELF_LOG"):
            src = enr["_SHELF_LOG"]
        elif self._get_exp_days() is not None:
            src = f"override_days={self._get_exp_days()}"
        self.logger.log(f"Срок годности: {src}; EXP_DATE={enr.get('EXP_DATE','')}")

        # источник ShortName
        self.logger.log(f"ShortName источник: {enr.get('_SHORT_SRC','?')} → '{enr.get('ShortName','')}'")
        return enr

    def _show_preview_path(self, path):
        try:
            img = Image.open(path)
            self.preview_label.update_idletasks()
            box_w = max(320, int(self.preview_label.winfo_width() or 560)) - 40
            box_h = max(240, int(self.preview_label.winfo_height() or 360)) - 40
            iw, ih = img.size
            ratio = min(box_w / iw, box_h / ih)
            new_w, new_h = max(1, int(iw * ratio)), max(1, int(ih * ratio))
            self.preview_ctkimg = ctk.CTkImage(light_image=img, dark_image=img, size=(new_w, new_h))
            self.preview_label.configure(image=self.preview_ctkimg, text="")
        except Exception as e:
            self.logger.err(f"Не удалось показать превью: {e}")
            self.preview_label.configure(text="Не удалось показать превью")

    def _maybe_calibrate(self, fmt, copies):
        """
        Новая логика: никаких отдельных калибровочных заданий.
        6 строк 'X' будут добавлены в enriched_rows/tmp_batch перед печатью каждого батча.
        Здесь только логируем.
        """
        try:
            if bool(self.calib_var.get()):
                self.logger.log("КАЛИБРОВКА: будет выполнена через tmp_batch (6 строк 'X').")
            else:
                self.logger.log("КАЛИБРОВКА: выключена.")
        except Exception:
            pass
        return

    
    def _print_marking_label(self, enr):
        if not AUTO_MARKING_ENABLED:
            return
        if not getattr(self, "bt", None):
            try:
                self.logger.err("Авто-маркировка: BarTender COM не инициализирован.")
            except Exception:
                pass
            return
        try:
            prn = (MARKING_PRINTER_NAME or "").strip()
        except Exception:
            prn = ""
        if not prn:
            try:
                self.logger.err("Авто-маркировка: принтер не задан (MARKING_PRINTER_NAME пуст).")
            except Exception:
                pass
            return
        template = MARKING_LABEL_TEMPLATE_PATH
        if not template:
            try:
                self.logger.err("Авто-маркировка: путь к шаблону не указан (MARKING_LABEL_TEMPLATE_PATH).")
            except Exception:
                pass
            return
        try:
            if not os.path.isabs(template):
                template = os.path.join(BASE_DIR, template)
        except Exception:
            pass
        if not os.path.isfile(template):
            try:
                self.logger.err(f"Авто-маркировка: шаблон не найден: {template}")
            except Exception:
                pass
            return
        fmt = None
        try:
            fmt = self.bt.open_format(template)
            self.bt.set_common_print_flags(fmt)
            try:
                fmt.PrintSetup.Printer = prn
            except Exception:
                pass
            try:
                fmt.PrintSetup.PrinterName = prn
            except Exception:
                pass
            try:
                fmt.PrintSetup.IdenticalCopiesOfLabel = 1
            except Exception:
                pass
            self.bt.apply_fields(fmt, enr)
            ok = self._bt_print(fmt, 1, False)
            if ok:
                try:
                    self.logger.log(f"Маркировочная этикетка отправлена → '{prn}'.")
                except Exception:
                    pass
            else:
                try:
                    self.logger.err(f"Авто-маркировка: печать не удалась на '{prn}'.")
                except Exception:
                    pass
        except Exception as e:
            try:
                self.logger.err(f"Авто-маркировка: сбой печати: {e}")
            except Exception:
                pass
        finally:
            try:
                fmt.Close(1)
            except Exception:
                pass


    def _bt_print(self, fmt, copies: int, show_dialog: bool):

        # Если выбран режим "Одно задание" — используем IdenticalCopiesOfLabel
        try:
            sj = False
            if hasattr(self, "single_job_var"):
                if hasattr(self.single_job_var, "get"):
                    sj = bool(self.single_job_var.get())
                else:
                    sj = bool(self.single_job_var)
            if sj and not show_dialog:
                try:
                    fmt.IdenticalCopiesOfLabel = int(copies)
                except Exception:
                    pass
                try:
                    self.logger.log(f"SingleJob: IdenticalCopiesOfLabel={copies}")
                except Exception:
                    pass
                try:
                    # Одно задание, без диалогов
                    fmt.PrintOut(False, True) if self._dialog_flag() else fmt.PrintOut(False, False)
                    return True
                except Exception as e:
                    try:
                        self.logger.err(f"SingleJob PrintOut error: {e}")
                    except Exception:
                        pass
                    # Падать не будем — ниже сработают обычные пути
        except Exception:
            pass
        if getattr(self, 'cancel_requested', False):
            try:
                self.logger.log('Печать отменена пользователем (до отправки).')
            except Exception:
                pass
            return False
        """
        Если show_dialog=True: пробуем 2-булевые перегрузки (A1/A2) и затем трёхаргументную (B).
        Если show_dialog=False: пропускаем A1/A2 полностью, используем B или тихий C.
        Ошибку 'Недопустимое число параметров' в B понижаем до инфо-лога, если show_dialog=False.
        """
        prompt = bool(show_dialog)

        if False and prompt:
            # A1: две булевые с ожиданием
            try:
                self.logger.log("BT Print: A1 -> PrintOut(True, True)")
                fmt.PrintOut(True, True)
                return True
            except Exception as e:
                self.logger.err(f"BT Print A1 ошибка: {e}")
            # A2: две булевые без ожидания
            try:
                self.logger.log("BT Print: A2 -> PrintOut(True, False)")
                fmt.PrintOut(True, False)
                return True
            except Exception as e:
                self.logger.err(f"BT Print A2 ошибка: {e}")

        # B: трёхаргументная перегрузка (copies, serialized, showDialog)
        try:
            self.logger.log(f"BT Print: B -> PrintOut(Copies={int(copies)}, Serialized=False, ShowDialog={prompt})")
            fmt.PrintOut(int(copies), False, prompt)
            return True
        except Exception as e:
            if False and prompt:
                self.logger.err(f"BT Print B ошибка: {e}")
            else:
                self.logger.log(f"BT Print: B недоступна (перехожу в C): {e}")

        # C: fallback — полностью тихо
        try:
            self.logger.log("BT Print: C -> PrintOut(False, False) (тихо)")
            fmt.PrintOut(False, True) if self._dialog_flag() else fmt.PrintOut(False, False)
            if False and prompt:
                self.logger.err("Диалог печати не поддерживается COM — выполнена тихая печать.")
            return True
        except Exception as e:
            self.logger.err(f"Печать не удалась: {e}")
            return False


    # ---------- pause helpers ----------
    def _toggle_pause(self):
        # Переключатель Пауза / Продолжить
        self.is_paused = not self.is_paused
        try:
            if getattr(self, "pause_btn", None):
                self.pause_btn.configure(text=("Продолжить" if self.is_paused else "Пауза"))
        except Exception:
            pass

    def _pause_wait(self):
        # Мягкая пауза: пока включена — поддерживаем отзывчивость GUI
        while self.is_paused:
            try:
                self.update_idletasks()
                self.update()
            except Exception:
                pass
            time.sleep(0.1)


    # ---------- date picker ----------
    def _open_date_picker(self):
        """
        Открывает простой выбор даты. Если установлен пакет tkcalendar — используем его.
        Иначе показываем диалог с полями ДД/ММ/ГГГГ.
        """
        try:
            import tkcalendar  # type: ignore
            top = ctk.CTkToplevel(self)
            top.title("Выбор даты производства")
            cal = tkcalendar.Calendar(top, selectmode='day', date_pattern='dd.mm.yyyy')
            cal.pack(padx=10, pady=10)
            def on_ok():
                self.prod_date_entry.delete(0, "end")
                self.prod_date_entry.insert(0, cal.get_date())
                top.destroy()
            ctk.CTkButton(top, text="OK", command=on_ok).pack(pady=8)
            ctk.CTkButton(top, text="Отмена", command=top.destroy).pack(pady=4)
            top.grab_set()
            return
        except Exception:
            pass

        # Фолбэк: простое окно с валидированными полями
        import datetime as _dt
        top = ctk.CTkToplevel(self)
        top.title("Выбор даты (ДД.ММ.ГГГГ)")
        frame = ctk.CTkFrame(top)
        frame.pack(padx=12, pady=12)

        # Текущая дата из поля
        cur = (self.prod_date_entry.get() or "").strip()
        try:
            d, m, y = [int(x) for x in cur.split(".")]
        except Exception:
            today = _dt.date.today()
            d, m, y = today.day, today.month, today.year

        day = ctk.CTkEntry(frame, width=40); day.insert(0, str(d).zfill(2)); day.pack(side="left", padx=4)
        ctk.CTkLabel(frame, text=".").pack(side="left")
        mon = ctk.CTkEntry(frame, width=40); mon.insert(0, str(m).zfill(2)); mon.pack(side="left", padx=4)
        ctk.CTkLabel(frame, text=".").pack(side="left")
        year = ctk.CTkEntry(frame, width=60); year.insert(0, str(y)); year.pack(side="left", padx=4)

        def only_digits(e):
            return re.fullmatch(r"[0-9]*", e or "") is not None
        vcmd = (self.register(only_digits), "%P")
        for w in (day, mon, year):
            w.configure(validate="key", validatecommand=vcmd)

        def on_ok():
            try:
                dd = int(day.get() or "0"); mm = int(mon.get() or "0"); yy = int(year.get() or "0")
                _dt.date(yy, mm, dd)  # проверка
                self.prod_date_entry.delete(0, "end")
                self.prod_date_entry.insert(0, f"{dd:02d}.{mm:02d}.{yy:04d}")
                top.destroy()
            except Exception:
                mb.showerror("Дата", "Введите корректную дату (ДД.ММ.ГГГГ).")
        btns = ctk.CTkFrame(top); btns.pack(pady=8)
        ctk.CTkButton(btns, text="OK", command=on_ok).pack(side="left", padx=6)
        ctk.CTkButton(btns, text="Отмена", command=top.destroy).pack(side="left", padx=6)
        top.grab_set()

# ---------- actions ----------

    def _preview(self):
        if not self.csv_path:
            mb.showerror("CSV", "Выбери CSV")
            return
        if not self.csv_rows:
            try:
                self.csv_rows = load_kontur_raw(self.csv_path)
            except Exception as e:
                self.logger.err(f"CSV ошибка: {e}")
                return
        if not self.csv_rows:
            mb.showerror("CSV", "Нет данных")
            return

        idx1 = min(self._get_index(), len(self.csv_rows))
        base = self.csv_rows[idx1 - 1]
        enr = self._enrich(base, idx1)
        if not enr:
            return

        src = "manual" if self.format_combo.get() in ("16x16", "30x20") else "auto"
        self.logger.log(f"Выбран формат: {enr['_FORMAT']} (источник: {src}, комбо: {self.format_combo.get()}, GTIN={base.get('GTIN','')})")
        fmt_name = enr["_FORMAT"]
        btw = self._get_btw_for_format(fmt_name)
        if not btw:
            return

        try:
            fmt = self.bt.open_format(btw)
            self.bt.set_common_print_flags(fmt)
            fmt.PrintSetup.IdenticalCopiesOfLabel = 1
            ok = self.bt.apply_fields(fmt, enr)
            if not ok:
                self.logger.err("Не удалось подставить значения (проверь имена полей в BTW).")

            out = os.path.join(os.path.dirname(btw), PREVIEW_NAME)
            # перезаписываем без вопросов
            try:
                if os.path.exists(out):
                    os.remove(out)
            except Exception:
                pass

            if self.bt.export_preview(fmt, out):
                self._show_preview_path(out)
                self.logger.log(f"Превью сохранено: {out} (формат {fmt_name})")
            else:
                self.logger.err("Не удалось сформировать превью.")
        except Exception as e:
            self.logger.err(f"Превью: {e}\n{traceback.format_exc()}")
        finally:
            try:
                fmt.Close(1)
            except Exception:
                pass

    def _print_one(self):
        self.cancel_requested = False
        prn = self._get_printer()
        if not prn:
            return
        if not self.csv_path:
            mb.showerror("CSV", "Выбери CSV")
            return
        if not self.csv_rows:
            try:
                self.csv_rows = load_kontur_raw(self.csv_path)
            except Exception as e:
                self.logger.err(f"CSV ошибка: {e}")
                return
        if not self.csv_rows:
            mb.showerror("CSV", "Нет данных")
            return

        idx1 = min(self._get_index(), len(self.csv_rows))
        copies = self._get_copies()
        base = self.csv_rows[idx1 - 1]
        enr = self._enrich(base, idx1)
        if not enr:
            return

        src = "manual" if self.format_combo.get() in ("16x16", "30x20") else "auto"
        self.logger.log(f"Выбран формат: {enr['_FORMAT']} (источник: {src}, комбо: {self.format_combo.get()}, GTIN={base.get('GTIN','')})")
        fmt_name = enr["_FORMAT"]
        btw = self._get_btw_for_format(fmt_name)
        if not btw:
            return

        self._pause_wait()
        self.logger.log(f"Печать строки #{idx1} (формат {fmt_name}), копий/шт={copies} → '{prn}'")
        try:
            fmt = self.bt.open_format(btw)
            self.bt.set_common_print_flags(fmt)
            fmt.PrintSetup.Printer = prn
            fmt.PrintSetup.IdenticalCopiesOfLabel = copies

            self._maybe_calibrate(fmt, copies)

            ok = self.bt.apply_fields(fmt, enr)
            if not ok:
                self.logger.err("Не удалось подставить значения (проверь имена полей в BTW).")

            main_ok = self._bt_print(fmt, fmt.PrintSetup.IdenticalCopiesOfLabel, False)
            if main_ok:
                self.logger.log(f"Основная этикетка отправлена → '{prn}'.")
                try:
                    self._print_marking_label(enr)
                except Exception:
                    try:
                        self.logger.err("Авто-маркировка: внутренняя ошибка вызова.")
                    except Exception:
                        pass
            else:
                self.logger.err("Основная этикетка не отправлена: ошибка печати.")
        except Exception as e:
            self.logger.err(f"Сбой печати 1 шт: {e}\n{traceback.format_exc()}")
        finally:
            try:
                fmt.Close(1)
            except Exception:
                pass

    def _print_all(self):

        try:
            self.deiconify(); self.state("normal"); self.lift()
        except Exception:
            pass
        self.cancel_requested = False
        prn = self._get_printer()

        if not prn:

            return

        if not self.csv_path:

            mb.showerror("CSV", "Выбери CSV")

            return

        if not self.csv_rows:

            try:

                self.csv_rows = load_kontur_raw(self.csv_path)

            except Exception as e:

                self.logger.err(f"CSV ошибка: {e}")

                return

        if not self.csv_rows:

            mb.showerror("CSV", "Нет данных")

            return

    

        limit = self._get_limit()

        idx0 = max(0, (self._get_index() if hasattr(self, '_get_index') else 1) - 1)
        rows_all = self.csv_rows[idx0:] if not limit else self.csv_rows[idx0:idx0+limit]

        total = len(rows_all)

        copies = self._get_copies()

        batch_size = self._get_batch_size() or total
        global_start = idx0
        self.logger.log(f"Старт со строки: {global_start+1}")

    

        # Сохраняем настройки

        try:

            self.cfg['batch_size'] = batch_size

            self.cfg['show_print_dialog'] = bool(self.show_dialog_var.get())

            save_config(self.cfg)

        except Exception:

            pass

    

        self.logger.log(f"Серия: {total}/{len(self.csv_rows)} строк; копий/шт={copies}; принтер='{prn}'; пакет={batch_size}; диалог={'ON' if self.show_dialog_var.get() else 'OFF'}")

        sent_total = 0

    

        batches = [(i, min(i+batch_size, total)) for i in range(0, total, batch_size)]

        self.logger.log(f"Всего пакетов: {len(batches)}")

    

        for bidx, (start, end) in enumerate(batches, start=1):

    

            # === PRECOMPUTE ENRICHED ROWS & WRITE tmp_batch.csv ===

    

            rows = rows_all[start:end]

    

            enriched_rows = []

    

            formats_in_batch = set()

    

            for i_pre, base_pre in enumerate(rows):

    

                self._pause_wait()

    

                try:

    

                    self.update_idletasks(); self.update()

    

                except Exception:

    

                    pass

    

                idx1_pre = global_start + start + i_pre + 1

    

                enr_pre = self._enrich(base_pre, idx1_pre)

    

                if not enr_pre:

    

                    self.logger.err(f"Строка {idx1_pre}: данные не сформированы — пропуск в буфере")

    

                    continue

    

                enriched_rows.append(enr_pre)

    

                formats_in_batch.add(enr_pre.get("_FORMAT",""))
            # === prepend 6 calibration rows in-memory (so they print as one job) ===
            try:
                if bool(self.calib_var.get()) and enriched_rows:
                    cols = getattr(self, 'REQ_COLS', ["ShortName","ShortGTIN","EXP_DATE","PROD_DATE","PART_NUM","DM","NUM"])
                    fmt0 = (enriched_rows[0].get("_FORMAT") or "16x16")
                    dummy = {k: ("1" if k.upper()=="NUM" else ("000" if k=="ShortGTIN" else "X")) for k in cols}
                    dummy.update({"_FORMAT": fmt0})
                    enriched_rows = [dummy.copy() for _ in range(6)] + enriched_rows
                    self.logger.log("КАЛИБРОВКА: 6 строк 'X' добавлены в начало батча.")
            except Exception:
                pass
            # Показ диалога — ровно один раз перед первым ярлыком батча
            prompt_left = bool(self.show_dialog_var.get())



    

            if enriched_rows:

    

                self._write_tmp_batch_csv(enriched_rows)

    

            else:

    

                self.logger.err(f"Пакет {bidx}: нет валидных строк для записи tmp_batch.csv — пропуск печати пакета")

    

                continue

    

            

    

            # === PRINT FROM enriched_rows ===

    

            fmt = None

    

            last_btw = None

    

            try:

    

                calib_done = False

    

                for i_enr, enr in enumerate(enriched_rows):

    

                    self._pause_wait()

    

                    try:

    

                        self.update_idletasks(); self.update()

    

                    except Exception:

    

                        pass

    

            

    

                    fmt_name = enr.get("_FORMAT","16x16")

    

                    btw = self._get_btw_for_format(fmt_name)

    

                    if not btw:

    

                        self.logger.err(f"Пакет {bidx}: BTW для формата {fmt_name} не указан — строка пропущена")

    

                        self._set_progress(i_enr+1, len(enriched_rows), f"Печать пакета {bidx}")

    

                        continue

    

            

    

                    if (fmt is None) or (btw != last_btw):

    

                        try:

    

                            if fmt:

    

                                fmt.Close(1)

    

                        except Exception:

    

                            pass

    

                        fmt = self.bt.open_format(btw)

    

                        self.bt.set_common_print_flags(fmt)

    

                        fmt.PrintSetup.Printer = prn

    

                        last_btw = btw

    

                        if not calib_done:

    

                            self._maybe_calibrate(fmt, copies)

    

                            calib_done = True

    

            

    

                    fmt.PrintSetup.IdenticalCopiesOfLabel = copies

    

                    ok = self.bt.apply_fields(fmt, enr)

    

                    if not ok:

    

                        self.logger.err(f"Пакет {bidx}: не удалось подставить значения — строка пропущена")

    

                        self._set_progress(i_enr+1, len(enriched_rows), f"Печать пакета {bidx}")

    

                        continue

    

            

    

                    main_ok = self._bt_print(fmt, fmt.PrintSetup.IdenticalCopiesOfLabel, prompt_left)

                    prompt_left = False

                    if main_ok:

                        sent_total += 1

                        try:
                            self.logger.log(f"Основная этикетка отправлена → '{prn}'.")
                        except Exception:
                            pass

                        try:
                            self._print_marking_label(enr)
                        except Exception:
                            try:
                                self.logger.err("Авто-маркировка: внутренняя ошибка вызова.")
                            except Exception:
                                pass

                        if (sent_total % 50) == 0:
                            self.logger.log(f"Отправлено: {sent_total}/{total}")

                    else:
                        self.logger.err("Основная этикетка не отправлена: ошибка печати.")

                    self._set_progress(i_enr+1, len(enriched_rows), f"Печать пакета {bidx}")

    

            except Exception as e:

    

                import traceback

    

                self.logger.err(f"Сбой печати пакета {bidx} (после записи tmp_batch): {e}\n{traceback.format_exc()}")

    

            finally:

    

                try:

    

                    if fmt:

    

                        fmt.Close(1)

    

                except Exception:

    

                    pass

    

            # === END NEW BLOCK ===

    

            # --- Confirm next batch ---

    

            if bidx < len(batches):

    

                _go_next = True

    

                try:

    

                    _go_next = mb.askyesno("Печать следующего пакета?", f"Печать пакета {bidx} завершена. Печатать следующий?")

    

                except Exception:

    

                    pass

    

                if not _go_next:

    

                    break


            prompt_left = bool(self.show_dialog_var.get())
            rows = rows_all[start:end]

            self.logger.log(f"[Пакет {bidx}/{len(batches)}] Строки {global_start+start+1}-{global_start+end} ({len(rows)} шт.)")

            self._set_progress(0, len(rows), f"Печать пакета {bidx}")

    

            fmt = None

            last_btw = None

    

            try:



    

                pass  # disabled original per-row loop
                for i, base in enumerate(rows):
                    self._pause_wait()
                    try:
                        self.update_idletasks()
                        self.update()
                    except Exception:
                        pass
                    idx1 = global_start + start + i + 1

                    enr = self._enrich(base, idx1)

                    if not enr:

                        self.logger.err(f"Строка {idx1}: данные не сформированы — пропуск")

                        self._set_progress(i+1, len(rows), f"Печать пакета {bidx}")

                        continue

    

                    src = "manual" if self.format_combo.get() in ("16x16", "30x20") else "auto"

                    self.logger.log(f"[P{bidx}] Строка {idx1}: формат {enr['_FORMAT']} (источник: {src}, комбо: {self.format_combo.get()}, GTIN={base.get('GTIN','')})")

    

                    fmt_name = enr["_FORMAT"]

                    btw = self._get_btw_for_format(fmt_name)

                    if not btw:

                        self.logger.err(f"Строка {idx1}: BTW для формата {fmt_name} не указан — пропуск")

                        self._set_progress(i+1, len(rows), f"Печать пакета {bidx}")

                        continue

    

                    if (fmt is None) or (btw != last_btw):

                        try:

                            if fmt:

                                fmt.Close(1)

                        except Exception:

                            pass

                        fmt = self.bt.open_format(btw)

                        self.bt.set_common_print_flags(fmt)

                        fmt.PrintSetup.Printer = prn

                        last_btw = btw

    

                        if not calib_done:

                            self._maybe_calibrate(fmt, copies)

                            calib_done = True

    

                    fmt.PrintSetup.IdenticalCopiesOfLabel = copies

                    ok = self.bt.apply_fields(fmt, enr)

                    if not ok:

                        self.logger.err(f"Строка {idx1}: не удалось подставить значения — пропуск")

                        self._set_progress(i+1, len(rows), f"Печать пакета {bidx}")

                        continue

    

                    main_ok = self._bt_print(fmt, fmt.PrintSetup.IdenticalCopiesOfLabel, prompt_left)
                    prompt_left = False

                    if main_ok:

                        sent_total += 1

                        try:
                            self.logger.log(f"Основная этикетка отправлена → '{prn}'.")
                        except Exception:
                            pass

                        try:
                            self._print_marking_label(enr)
                        except Exception:
                            try:
                                self.logger.err("Авто-маркировка: внутренняя ошибка вызова.")
                            except Exception:
                                pass

                        if (sent_total % 50) == 0:

                            self.logger.log(f"Отправлено: {sent_total}/{total}")

                    else:

                        self.logger.err("Основная этикетка не отправлена: ошибка печати.")

    

                    self._set_progress(i+1, len(rows), f"Печать пакета {bidx}")

    

            except Exception as e:

                import traceback

                self.logger.err(f"Сбой печати пакета {bidx}: {e}\n{traceback.format_exc()}")

            finally:

                try:

                    if fmt:

                        fmt.Close(1)

                except Exception:

                    pass

    

            if bidx < len(batches):

                cont = mb.askyesno("Продолжить печать?", f"Пакет {bidx} завершён. Печатать следующий пакет ({bidx+1}/{len(batches)})?")

                if not cont:

                    self.logger.log("Печать остановлена по запросу пользователя.")

                    break

    

        self._set_progress(total, total, "Печать")

        self.logger.log(f"Готово. Отправлено: {sent_total}/{total}")
        # ## GUI_VIS_REVEAL_END ##
        try:
            self.deiconify(); self.state("normal"); self.lift()
        except Exception:
            pass


    

    def _print_test(self):
        """Печать одной тестовой страницы (лучше калибровки 'X'). 
        Если есть C:\\auto_print\\test_page.btw — используем его.
        Иначе — печатаем текущим выбранным шаблоном.
        """
        prn = self._get_printer()
        if not prn:
            return

        test_btw = os.path.join(BASE_DIR, "test_page.btw")
        if os.path.exists(test_btw):
            btw = test_btw
        else:
            # fallback: используем текущий btw
            fmt_name = "16x16" if self.format_combo.get() == "16x16" else (
                "30x20" if self.format_combo.get() == "30x20" else "16x16"
            )
            btw = self._get_btw_for_format(fmt_name)
            if not btw:
                return

        self.logger.log(f"Тестовая страница → '{prn}' (BTW: {btw})")
        try:
            fmt = self.bt.open_format(btw)
            self.bt.set_common_print_flags(fmt)
            fmt.PrintSetup.Printer = prn
            fmt.PrintSetup.IdenticalCopiesOfLabel = 1

            # безопасные значения (если в шаблоне нет полей — тихо пропустим)
            test_payload = {
                "DM": "TEST",
                "ShortName": "TEST NAME",
                "ShortGTIN": "000",
                "PROD_DATE": "01.01.2025",
                "EXP_DATE": "01.01.2026",
                "PART_NUM": "TEST",
                "NUM": "1"
            }
            self.bt.apply_fields(fmt, test_payload)

            self._bt_print(fmt, fmt.PrintSetup.IdenticalCopiesOfLabel, False)
            self.logger.log("Тестовая страница отправлена.")
        except Exception as e:
            self.logger.err(f"ERROR: Тестовая печать: {e}\n{traceback.format_exc()}")
        finally:
            try:
                fmt.Close(1)
            except Exception:
                pass

    # ---------- save log ----------

    def _save_log(self):
        data = self.logbox.get("1.0", "end-1c")
        if not data.strip():
            mb.showinfo("Сохранение лога", "Журнал пуст.")
            return
        path = fd.asksaveasfilename(title="Сохранить журнал", defaultextension=".txt", filetypes=[("Text", "*.txt")])
        if not path:
            return
        with open(path, "w", encoding="utf-8") as f:
            f.write(data)
        mb.showinfo("Сохранение лога", f"Сохранено:\n{path}")

    def _cancel_print(self):
        """Пользовательская отмена печати: ставит флаг, сбрасывает паузу и чистит очередь."""
        try:
            self.cancel_requested = True
            if getattr(self, "pause_btn", None):
                try:
                    self.is_paused = False
                    self.pause_btn.configure(text="Пауза")
                except Exception:
                    pass
            if getattr(self, "logger", None):
                self.logger.log("Отмена печати: флаг установлен. Пытаюсь очистить очередь...")
        except Exception:
            pass
        try:
            prn = self._get_printer()
        except Exception:
            prn = None
        if prn:
            try:
                self._purge_printer_queue(prn)
            except Exception as e:
                try:
                    self.logger.err(f"Очистка очереди не удалась: {e}")
                except Exception:
                    pass


    def _purge_printer_queue(self, prn_name):
        """Очистить очередь печати средствами Windows (PowerShell)."""
        import subprocess
        try:
            ps = f'Get-PrintJob -PrinterName "{prn_name}" -ErrorAction SilentlyContinue | Remove-PrintJob -Confirm:$false'
            cmd = ["powershell", "-NoProfile", "-Command", ps]
            subprocess.run(cmd, check=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            if getattr(self, "logger", None):
                self.logger.log(f"Очередь принтера '{prn_name}' очищена (если были задания).")
        except Exception as e:
            try:
                subprocess.run(["powershell","-NoProfile","-Command","Stop-Service Spooler -Force"], check=False)
                subprocess.run(["powershell","-NoProfile","-Command","Start-Service Spooler"], check=False)
                if getattr(self, "logger", None):
                    self.logger.log("Служба печати перезапущена.")
            except Exception:
                raise e



    def _print_one_single_job(self):


        # Печать выбранной строки N копиями: одно Windows-задание


        prn = self._get_printer()


        if not prn:


            return


        if not self.csv_path:


            mb.showerror("CSV", "Выбери CSV")


            return


        try:


            idx = self._get_index()


            copies = self._get_copies()


        except Exception:


            mb.showerror("Печать", "Некорректные индекс/копии")


            return


        if not self.csv_rows:


            try:


                self.csv_rows = load_kontur_raw(self.csv_path)


            except Exception as e:


                self.logger.err(f"CSV ошибка: {e}")


                return


        if not self.csv_rows:


            mb.showerror("CSV", "Нет данных")


            return



        base = self.csv_rows[max(1, min(idx, len(self.csv_rows))) - 1]


        enr = self._enrich(base, idx)


        if not enr:


            return



        fmt_name = enr['_FORMAT']


        path = self._get_btw_for_format(fmt_name)


        if not path:


            mb.showerror("BTW", f"Нет шаблона для формата {fmt_name}")


            return



        fmt = self._prepare_btw(path, enr, prn)


        if not fmt:


            return



        # Назначим принтер и копии во всех местах


        try:


            fmt.PrintSetup.PrinterName = prn


        except Exception:


            pass


        try:


            fmt.PrintSetup.Printer = prn


        except Exception:


            pass


        try:


            fmt.PrintSetup.IdenticalCopiesOfLabel = int(copies)


        except Exception:


            pass


        try:


            fmt.IdenticalCopiesOfLabel = int(copies)


        except Exception:


            pass



        # Диалог или тихо


        prompt = False


        try:


            prompt = bool(self.show_dialog_var.get())


        except Exception:


            prompt = False



        try:


            if False and prompt:


                try:


                    self.logger.log(f"Printer='{prn}', Copies={copies} → диалог")


                except Exception:


                    pass


                fmt.PrintOut(True, True)


                self.logger.log("Отправлено одно задание через диалог (копии задаются в окне).")


            else:


                try:


                    self.logger.log(f"Printer='{prn}', Copies={copies} → тихо")


                except Exception:


                    pass


                # одно Windows-задание, N одинаковых копий


                fmt.PrintOut(False, True) if self._dialog_flag() else fmt.PrintOut(False, False)


                self.logger.log(f"Отправлено одно задание на {copies} шт.")


        except Exception as ex:


            self.logger.err(f"SingleJob PrintOut error: {ex}")




    def _print_one_pdf_dialog(self):



        # Печать выбранной строки в PDF через системный диалог принтера



        prn = self._get_printer()



        if not prn:



            mb.showerror("Принтер", "Выбери принтер (можно выбрать 'Microsoft Print to PDF')")



            return



        if not self.csv_path:



            mb.showerror("CSV", "Выбери CSV")



            return



        try:



            idx = self._get_index()



            copies = self._get_copies()



        except Exception:



            mb.showerror("Печать", "Некорректные индекс/копии")



            return



        if not self.csv_rows:



            try:



                self.csv_rows = load_kontur_raw(self.csv_path)



            except Exception as e:



                self.logger.err(f"CSV ошибка: {e}")



                return



        if not self.csv_rows:



            mb.showerror("CSV", "Нет данных")



            return



        base = self.csv_rows[max(1, min(idx, len(self.csv_rows))) - 1]



        enr = self._enrich(base, idx)



        if not enr:



            return



        fmt_name = enr['_FORMAT']



        path = self._get_btw_for_format(fmt_name)



        if not path:



            mb.showerror("BTW", f"Нет шаблона для формата {fmt_name}")



            return



        fmt = self._prepare_btw(path, enr, prn)



        if not fmt:



            return




        # Зафиксируем копии и принтер



        try:



            fmt.PrintSetup.PrinterName = prn



        except Exception:



            pass



        try:



            fmt.PrintSetup.Printer = prn



        except Exception:



            pass



        try:



            fmt.PrintSetup.IdenticalCopiesOfLabel = int(copies)



        except Exception:



            pass



        try:



            fmt.IdenticalCopiesOfLabel = int(copies)



        except Exception:



            pass




        try:



            self.logger.log(f"Printer='{prn}', Copies={copies} → PDF dialog")



        except Exception:



            pass




        try:



            fmt.PrintOut(True, True)  # показать диалог, дождаться



            self.logger.log(f"Задание на PDF отправлено через диалог, копий: {copies}")



        except Exception as e:



            self.logger.err(f"PDF PrintOut error: {e}")




    def _prepare_btw(self, fmt_or_path, enr, prn_name):


        """


        Подготовка BTW:


        - гарантируем BarTender COM (self.app)


        - если на входе строка пути -> открываем через self.app.Formats.Open(...)


        - назначаем принтер (оба свойства, для совместимости)


        - подставляем NamedSubStrings из enr


        Возвращает объект Format или None.


        """


        try:


            # 0) Инициализация BarTender COM, если не создан


            bt = getattr(self, "app", None)


            if bt is None:


                from win32com.client import Dispatch


                self.app = Dispatch("BarTender.Application")


                try:


                    self.app.Visible = True


                except Exception:


                    pass


                bt = self.app



            # 1) Открыть BTW если передан путь


            fmt = fmt_or_path


            if isinstance(fmt_or_path, str):


                path = fmt_or_path
            try:
                import os
                # нормализуем на Windows
                norm = os.path.normpath(path)
                path = norm.replace('/', '\\')
                try:
                    self.logger.log(f"Нормализованный путь: {path}")
                except Exception:
                    pass
                if not os.path.isfile(path):
                    try:
                        self.logger.err(f"BTW файл не найден на диске: {path}")
                    except Exception:
                        pass
            except Exception:
                pass


                try:


                    self.logger.log(f"Открытие шаблона: {path}")


                except Exception:


                    pass


                try:


                    fmt = self.app.Formats.Open(path, False, "")


                except Exception as e:


                    try:


                        self.logger.err(f"Не удалось открыть BTW: {e}")


                    except Exception:


                        pass


                    return None



            # 2) Назначить принтер


            try:


                fmt.PrintSetup.PrinterName = prn_name


            except Exception:


                pass


            try:


                fmt.PrintSetup.Printer = prn_name


            except Exception:


                pass



            # 2.1) Безопасные флаги


            for a, v in (("UseDatabase", False), ("SelectRecordsAtPrint", False)):


                try:


                    setattr(fmt, a, v)


                except Exception:


                    pass



            # 3) Подставить NamedSubStrings


            try:


                names = [ss.Name for ss in fmt.NamedSubStrings]


                try:


                    self.logger.log(f"NamedSubStrings: {names}")


                except Exception:


                    pass


                set_ok = 0


                for name in names:


                    if name in enr:


                        try:


                            fmt.NamedSubStrings[name].Value = str(enr[name])


                            set_ok += 1


                        except Exception:


                            pass


                missed = [k for k in ['GTIN','NAME','ShortName','ShortGTIN','EXP_DATE','PROD_DATE','PART_NUM','DM','NUM'] if k not in names]


                try:


                    if missed:


                        self.logger.log(f"Подстановка: пропущены поля (нет в шаблоне): {missed}")


                    self.logger.log(f"Подстановка полей: всего={len(names)}, успешно={set_ok}")


                except Exception:


                    pass


            except Exception:


                pass



            return fmt


        except Exception as e:


            import traceback


            try:


                self.logger.err(f"_prepare_btw error: {e}\n{traceback.format_exc()}")


            except Exception:


                pass


            return None
# ------------------------ main ------------------------

        # === RANGE→ONE JOB через CSV (минимальный блок) ===
        def _rangecsv_only_digits(self, s):
            import re
            return re.sub(r"\D+", "", s or "")

        def _rangecsv_short_gtin(self, gtin):
            return self._rangecsv_only_digits(gtin).lstrip("0")

        def _rangecsv_parse_ru_date(self, txt):
            import datetime as dt
            txt = (txt or "").strip()
            return dt.datetime.strptime(txt, "%d.%m.%Y").date()

        def _rangecsv_add_days(self, d, n):
            import datetime as dt
            return d + dt.timedelta(days=int(n))

        def _rangecsv_add_months(self, d, months):
            from calendar import monthrange
            y = d.year + (d.month - 1 + months) // 12
            m = (d.month - 1 + months) % 12 + 1
            day = min(d.day, monthrange(y, m)[1])
            import datetime as dt
            return dt.date(y, m, day)

        def _rangecsv_calc_exp(self, prod_date, shelf_dict, override_days):
            if override_days is not None:
                return self._rangecsv_add_days(prod_date, int(override_days))
            if not shelf_dict:
                return None
            if shelf_dict.get("years"):
                return self._rangecsv_add_months(prod_date, 12*int(shelf_dict["years"]))
            if shelf_dict.get("months"):
                return self._rangecsv_add_months(prod_date, int(shelf_dict["months"]))
            if shelf_dict.get("weeks"):
                return self._rangecsv_add_days(prod_date, 7*int(shelf_dict["weeks"]))
            if shelf_dict.get("days"):
                return self._rangecsv_add_days(prod_date, int(shelf_dict["days"]))
            return None

        def _rangecsv_make_part(self, prod_date, part_tpl, manual, auto_flag):
            yymmdd = prod_date.strftime("%y%m%d")
            if not auto_flag and manual:
                return manual
            if part_tpl:
                return part_tpl.replace("{DATE}", yymmdd)
            return yymmdd

        def _rangecsv_enrich_row(self, base_row: dict, idx1: int):
            dm   = base_row.get("DM","")
            gtin = base_row.get("GTIN","")
            name = base_row.get("NAME","")

            try:
                pd = self._rangecsv_parse_ru_date(self.prod_date_entry.get())
            except Exception as e:
                from tkinter import messagebox as mb
                mb.showerror("Дата производства", f"Неверный формат (ДД.ММ.ГГГГ): {e}")
                return None

            info   = getattr(self, "product_map", {}).get(self._rangecsv_only_digits(gtin), {})
            shelf  = info.get("SHELF") or {}
            short_from_xlsx = (info.get("SHORTNAME") or "").strip()
            part_tpl = info.get("PART_TEMPLATE") or ""

            exp_override = None
            try:
                t = (self.exp_days_entry.get() or "").strip()
                if t:
                    exp_override = int(t)
            except Exception:
                exp_override = None

            exp_date = self._rangecsv_calc_exp(pd, shelf, exp_override)
            part_num = self._rangecsv_make_part(pd, part_tpl, (self.part_entry.get() or "").strip(), self.part_auto_var.get())

            return {
                "ShortName": (short_from_xlsx if short_from_xlsx else name)[:50],
                "ShortGTIN": self._rangecsv_short_gtin(gtin),
                "EXP_DATE":  exp_date.strftime("%d.%m.%Y") if exp_date else "",
                "PROD_DATE": pd.strftime("%d.%m.%Y"),
                "PART_NUM":  part_num,
                "DM":        dm,
                "NUM":       str(idx1),
            }

        def _rangecsv_write_tmp(self, path, rows, sep=";"):
            import csv
            self.logger.log(f"[tmp_csv] Поля: {self.REQ_COLS}")
            self.logger.log(f"[tmp_csv] Строк для записи: {len(rows)}")
            try:
                for i, r in enumerate(rows[:3]):
                    self.logger.log(f"[tmp_csv] Превью {i+1}: " + ", ".join(str(r.get(k, "")) for k in self.REQ_COLS))
            except Exception as e:
                self.logger.err(f"[tmp_csv] Ошибка превью: {e}")

            with open(path, "w", encoding="utf-8", newline="") as f:
                w = csv.DictWriter(f, fieldnames=self.REQ_COLS, delimiter=sep, quoting=csv.QUOTE_MINIMAL)
                w.writeheader()
                for r in rows:
                    w.writerow({k: r.get(k, "") for k in self.REQ_COLS})

        def _rangecsv_repoint_db(self, fmt, csv_path):
            changed = False
            try:
                dbs = fmt.DatabaseConnections
                for i in range(1, dbs.Count + 1):
                    dc = dbs.Item(i)
                    for attr in ("TextFile", "TextFileInfo", "ODBC", "OLEDB"):
                        try:
                            obj = getattr(dc, attr)
                        except Exception:
                            continue
                        for prop in ("Filename", "FileName", "Name", "Path", "ConnectionString"):
                            try:
                                _ = getattr(obj, prop)
                            except Exception:
                                continue
                            try:
                                setattr(obj, prop, csv_path)
                                changed = True
                            except Exception:
                                pass
            except Exception:
                pass
            try:
                fmt.UseDatabase = True
                fmt.SelectRecordsAtPrint = False
            except Exception:
                pass
            return changed

        def _rangecsv_choose_format_for_row(self, base_row):
            choice = self.format_combo.get() if hasattr(self, "format_combo") else "Auto"
            return "16x16" if choice == "Auto" else choice

        def _rangecsv_get_btw(self, fmt_name: str):
            import os
            path = self.btw16_entry.get().strip() if fmt_name == "16x16" else self.btw30_entry.get().strip()
            if not path or not os.path.isfile(path):
                from tkinter import messagebox as mb
                mb.showerror("BTW", f"Укажи BTW для {fmt_name}")
                return None
            return path


        def _print_range_one_job_via_csv(self):

                import csv, time, os, traceback


                # ------- параметры прогресса --------

                LOG_EVERY = 100         # писать в лог каждые N строк

                PROGRESS_EVERY = 50     # шаг обновления прогресса


                self.cancel_requested = False


                # Проверяем входные данные

                if not getattr(self, "csv_path", None):

                    try:

                        from tkinter import messagebox as mb

                        mb.showerror("CSV", "Выбери CSV")

                    except Exception:

                        pass

                    return

                if not getattr(self, "csv_rows", None):

                    try:

                        self.csv_rows = load_kontur_raw(self.csv_path)

                    except Exception as e:

                        self.logger.err(f"CSV ошибка: {e}")

                        return

                if not self.csv_rows:

                    try:

                        from tkinter import messagebox as mb

                        mb.showerror("CSV", "Нет данных")

                    except Exception:

                        pass

                    return


                # Диапазон

                idx0 = max(0, self._get_index() - 1)

                limit = self._get_limit()

                rows_all = self.csv_rows[idx0:] if not limit else self.csv_rows[idx0:idx0+limit]

                total = len(rows_all)


                # Подготовка печати

                prn = self._get_printer()

                if not prn:

                    return

                copies = self._get_copies()

                try:

                    show_dialog = bool(self.show_dialog_var.get())

                except Exception:

                    show_dialog = False


                # Обогащение всех строк в память

                t0 = time.time()

                enriched = []

                self.logger.log("[INFO] Сбор данных в память для tmp_batch.csv (ускоренный режим)")

                for i, base in enumerate(rows_all):

                    if self.cancel_requested:

                        self.logger.log("Отменено пользователем во время подготовки CSV.")

                        return

                    idx1 = idx0 + i + 1

                    enr = self._enrich(base, idx1)

                    if not enr:

                        self.logger.err(f"[WARN] Строка {idx1}: данные не сформированы — пропуск")

                        continue

                    enriched.append(enr)

                    if (i+1) % LOG_EVERY == 0:

                        self.logger.log(f"[PROGRESS] Подготовлено {i+1}/{total}")

                    if (i+1) % PROGRESS_EVERY == 0:

                        self._set_progress(i+1, total, "Подготовка CSV")


                if not enriched:

                    self.logger.err("Нет валидных строк для печати.")

                    return


                # Запись CSV одним махом

                tmp_path = os.path.join(os.path.dirname(self.csv_path) if os.path.isdir(os.path.dirname(self.csv_path)) else os.getcwd(), "tmp_batch.csv")

                try:

                    with open(tmp_path, "w", encoding="utf-8", newline="") as f:

                        w = csv.writer(f, delimiter=";")

                        w.writerow(self.REQ_COLS)

                        for i, enr in enumerate(enriched, 1):

                            w.writerow([

                                enr.get("ShortName",""),

                                enr.get("ShortGTIN",""),

                                enr.get("EXP_DATE",""),

                                enr.get("PROD_DATE",""),

                                enr.get("PART_NUM",""),

                                enr.get("DM",""),

                                enr.get("NUM",""),

                            ])

                except Exception as e:

                    self.logger.err(f"Не удалось записать tmp_batch.csv: {e}")

                    return


                elapsed = round(time.time() - t0, 3)

                self.logger.log(f"[INFO] CSV сформирован: {len(enriched)} строк → {tmp_path} (время: {elapsed} сек)")


                # Печать одним заданием через уже привязанный к шаблону текстовый источник

                fmt_name = enriched[0].get("_FORMAT", "16x16")

                btw = self._get_btw_for_format(fmt_name)

                if not btw:

                    return


                try:

                    fmt = self.bt.open_format(btw)

                except Exception as e:

                    self.logger.err(f"Не удалось открыть BTW: {e}")

                    return


                try:

                    # Лог по DatabaseConnections

                    try:

                        dbconns = getattr(fmt, "DatabaseConnections", None)

                        count = int(dbconns.Count) if dbconns else 0

                    except Exception:

                        dbconns = None

                        count = 0

                    self.logger.log(f"[DB] DatabaseConnections.Count = {count}")

                    if count >= 1 and dbconns:

                        try:

                            dbconns.Item(1).DatabaseFilename = tmp_path

                            self.logger.log(f"[DB] Установлен DatabaseFilename = {tmp_path}")

                        except Exception as e:

                            self.logger.err(f"[DB] Не удалось установить DatabaseFilename: {e}")

                    else:

                        self.logger.err("В шаблоне нет DatabaseConnections — печать будет попытана как обычная.")


                    # Принтер и копии

                    try: fmt.PrintSetup.Printer = prn

                    except Exception: pass

                    try: fmt.PrintSetup.PrinterName = prn

                    except Exception: pass

                    try: fmt.PrintSetup.IdenticalCopiesOfLabel = int(copies)

                    except Exception: pass


                    # Отправка

                    if getattr(self, "single_job_var", None):

                        try:

                            sj = bool(self.single_job_var.get())

                        except Exception:

                            sj = bool(self.single_job_var)

                    else:

                        sj = True


                    if sj:

                        self.logger.log("[INFO] Отправка одним заданием (CSV)")

                        try:

                            fmt.PrintOut(int(copies), False, False)

                        except Exception:

                            fmt.PrintOut(False, True) if self._dialog_flag() else fmt.PrintOut(False, False)

                    else:

                        fmt.PrintOut(False, True) if self._dialog_flag() else fmt.PrintOut(False, False)


                    self.logger.log("Готово: пакет отправлен.")

                except Exception as e:

                    self.logger.err(f"Сбой печати из CSV: {e}\n{traceback.format_exc()}")

                finally:

                    try:

                        fmt.Close(1)

                    except Exception:

                        pass

def main():
    app = App()
    app.mainloop()

# === begin MONKEY PATCH (range one job + pdf dialog) ===

def _ap_write_tmp_batch_csv(self, tmp_csv_path, rows, req_cols=None):
    import os, csv

    req = req_cols or ["ShortName","ShortGTIN","EXP_DATE","PROD_DATE","PART_NUM","DM","NUM"]

    def _cal_enabled_any(self):
        # Scan any attribute resembling calibration flag
        try:
            for name, val in getattr(self, "__dict__", {}).items():
                lname = str(name).lower()
                if ("cal" in lname) or ("калибр" in lname):
                    try:
                        v = val.get() if hasattr(val, "get") else val
                        if bool(v):
                            return True
                    except Exception:
                        pass
        except Exception:
            pass
        # explicit fallbacks
        for _n in ("calib_var","calibrate_var","calibration_var","calib_check","calibrate_check","calib_chk"):
            try:
                v = getattr(self, _n, None)
                if v is None: 
                    continue
                v = v.get() if hasattr(v, "get") else v
                if bool(v):
                    return True
            except Exception:
                pass
        return False

    do_cal = _cal_enabled_any(self)

    rows2 = list(rows or [])

    # prepend 6 X rows (dedup) when enabled
    if do_cal and rows2:
        def _is_x_row(r: dict) -> bool:
            try:
                for k in req:
                    v = str(r.get(k, "") or "")
                    if k.upper() == "NUM":
                        if v not in ("", "0", "1"):
                            return False
                    elif k == "ShortGTIN":
                        if v not in ("", "0", "000"):
                            return False
                    else:
                        if v != "X":
                            return False
                return True
            except Exception:
                return False
        first = rows2[0]
        if not _is_x_row(first):
            dummy = {k: ("1" if k.upper()=="NUM" else ("000" if k=="ShortGTIN" else "X")) for k in req}
            try:
                fmt0 = first.get("_FORMAT") or "16x16"
                dummy["_FORMAT"] = fmt0
            except Exception:
                pass
            rows2 = [dummy.copy() for _ in range(6)] + rows2
            try: self.logger.log("CSV-CAL: добавил 6 'X' в начало tmp_batch.csv")
            except Exception: pass
    else:
        try: self.logger.log(f"CSV-CAL: флаг={do_cal} — вставка 'X' не выполнялась")
        except Exception: pass

    # write CSV (comma + quotes, UTF-8-SIG)
    try:
        os.makedirs(os.path.dirname(tmp_csv_path), exist_ok=True)
    except Exception:
        pass

    with open(tmp_csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=req, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)
        w.writeheader()
        for r in rows2:
            w.writerow({k: (r.get(k, "") or "") for k in req})

    try:
        self.logger.log(f"tmp_batch.csv записан: {tmp_csv_path} (строк={len(rows2)})")
    except Exception:
        pass

    return tmp_csv_path
def _ap_repoint_text_db(self, fmt, csv_path):
    # Перепривязывает первое подключение БД в шаблоне к CSV (если подключение есть).
    try:
        conns = getattr(fmt, "DatabaseConnections", None)
        if not conns:
            try: self.logger.err("В шаблоне нет DatabaseConnections — печать из CSV одним заданием недоступна.")
            except Exception: pass
            return False
        ok = False
        for conn in conns:
            for attr in ("TextFile","AdoTextFile","ODBC","OLEDB"):
                try:
                    tf = getattr(conn, attr, None)
                except Exception:
                    tf = None
                if not tf:
                    continue
                # путь к файлу
                for prop in ("FileName","FileNameFull","Filename","Name","Path"):
                    try:
                        setattr(tf, prop, csv_path)
                        ok = True
                        break
                    except Exception:
                        continue
                # разделители и заголовок
                for k, v in (("HasFieldNames", True),
                             ("HasHeaderRecord", True),
                             ("FieldDelimiter", ","),
                             ("Delimiter", ","),
                             ("RecordDelimiter", "\n"),
                             ("UseDoubleQuotes", True)):
                    try:
                        setattr(tf, k, v)
                    except Exception:
                        pass
            if ok:
                break
        for a, v in (("UseDatabase", True), ("SelectRecordsAtPrint", False), ("RecordRange", "1")):
            try: setattr(fmt, a, v)
            except Exception: pass
        try:
            self.logger.log(f"DB repoint: {'OK' if ok else 'NO'}; UseDatabase={getattr(fmt,'UseDatabase',None)}")
        except Exception:
            pass
        return ok
    except Exception as e:
        try:
            import traceback
            self.logger.err(f"repoint_text_db error: {e}\n{traceback.format_exc()}")
        except Exception:
            pass
        return False

def _ap_open_fmt(self, btw_path):
    # Универсально открыть BTW либо через self.bt, либо прямым Dispatch
    try:
        if hasattr(self, "bt") and getattr(self.bt, "open_format", None):
            return self.bt.open_format(btw_path)
    except Exception:
        pass
    try:
        from win32com.client import Dispatch
        app = getattr(self, "app", None)
        if app is None:
            app = Dispatch("BarTender.Application")
            self.app = app
        return app.Formats.Open(btw_path, False, "")
    except Exception as e:
        try: self.logger.err(f"Не удалось открыть BTW: {e}")
        except Exception: pass
        return None

def _print_one_pdf_dialog_patched(self):
    prn = self._get_printer()
    if not prn: return
    if not getattr(self, "csv_path", ""):
        from tkinter import messagebox as mb
        mb.showerror("CSV", "Выбери CSV"); return
    if not getattr(self, "csv_rows", []):
        try: self.csv_rows = load_kontur_raw(self.csv_path)
        except Exception as e:
            self.logger.err(f"CSV ошибка: {e}"); return
    idx1 = min(self._get_index(), len(self.csv_rows))
    base = self.csv_rows[idx1-1]
    enr = self._enrich(base, idx1)
    if not enr: return

    fmt_name = enr.get("_FORMAT") or (self.format_combo.get() if hasattr(self, "format_combo") else "16x16")
    btw = self._get_btw_for_format(fmt_name)
    if not btw: return
    fmt = _ap_open_fmt(self, btw)
    if not fmt: return
    try:
        try: self.bt.set_common_print_flags(fmt)
        except Exception: pass
        try: fmt.PrintSetup.Printer = prn
        except Exception: pass
        try: fmt.PrintSetup.IdenticalCopiesOfLabel = 1
        except Exception: pass
        try: self.bt.apply_fields(fmt, enr)
        except Exception: pass
        try: self.logger.log(f"Printer='{prn}' → диалог")
        except Exception: pass
        fmt.PrintOut(True, True)
        self.logger.log("Задание отправлено через диалог.")
    except Exception as e:
        self.logger.err(f"PDF PrintOut error: {e}")
    finally:
        try: fmt.Close(1)
        except Exception: pass

def _print_range_one_job_via_csv_patched(self):
    import os
    prn = self._get_printer()
    if not prn: return
    if not getattr(self, "csv_path", ""):
        from tkinter import messagebox as mb
        mb.showerror("CSV", "Выбери CSV"); return
    if not getattr(self, "csv_rows", []):
        try: self.csv_rows = load_kontur_raw(self.csv_path)
        except Exception as e:
            self.logger.err(f"CSV ошибка: {e}"); return
    total = len(self.csv_rows)
    if total == 0:
        from tkinter import messagebox as mb
        mb.showerror("CSV", "Нет данных"); return

    try:
        import tkinter.simpledialog as sd
        count = sd.askinteger("Сколько штук?", "Введите количество этикеток (шт.):", minvalue=1, initialvalue=10, parent=self)
    except Exception:
        count = 10
    if not count: return
    start1 = self._get_index()
    end1 = min(total, start1 + count - 1)

    out_rows = []
    fmt_name_first = None
    for idx1 in range(start1, end1+1):
        base = self.csv_rows[idx1-1]
        enr = self._enrich(base, idx1)
        if not enr:
            continue
        if fmt_name_first is None:
            fmt_name_first = enr.get("_FORMAT")
        out_rows.append(enr)
    if not out_rows:
        self.logger.err("Нечего печатать (после обогащения пусто)."); return

    btw = self._get_btw_for_format(fmt_name_first or "16x16")
    if not btw: return

    tmp_csv = os.path.join(os.path.dirname(self.csv_path), "tmp_batch.csv")
    _ap_write_tmp_batch_csv(self, tmp_csv, out_rows, getattr(self, "REQ_COLS", None))

    fmt = _ap_open_fmt(self, btw)
    if not fmt: return
    try:
        try: fmt.PrintSetup.Printer = prn
        except Exception: pass
        ok_db = _ap_repoint_text_db(self, fmt, tmp_csv)
        if not ok_db:
            self.logger.err("Не удалось перепривязать БД к tmp_batch.csv — проверь, что BTW привязан к текстовой БД.")
            return
        try: self.logger.log(f"ONE-JOB: {start1}-{end1} → диалог печати")
        except Exception: pass
        fmt.PrintOut(True, True)
        self.logger.log("ONE-JOB отправлено.")
    except Exception as e:
        self.logger.err(f"ONE-JOB error: {e}")
    finally:
        try: fmt.Close(1)
        except Exception: pass

# Привязываем методы к классу до вызова main()
try:
    App._print_one_pdf_dialog = _print_one_pdf_dialog_patched
    App._print_range_one_job_via_csv = _print_range_one_job_via_csv_patched
    if not hasattr(App, "REQ_COLS"):
        App.REQ_COLS = ["ShortName","ShortGTIN","EXP_DATE","PROD_DATE","PART_NUM","DM","NUM"]
except Exception as _e:
    # лог в stdout — чтобы не рушить импорт
    print("Monkey-patch warning:", _e)

# === end MONKEY PATCH ===

# === Monkey patches: DB log + tmp_batch regenerate, no-rebind (BEGIN) ===
import csv, datetime as _dt, os as _os, traceback as _tb, re as _re


# --- Module-level SAFE calibration helper (defined early) ---
def _calibrate_fmt_safely(self, fmt):
    """
    Печатает 6 тестовых этикеток 'X' без БД и с копиями=1, затем восстанавливает состояние.
    """
    # Сохранение состояния
    try: prev_use_db = getattr(fmt, "UseDatabase", None)
    except Exception: prev_use_db = None
    try: prev_sel = getattr(fmt, "SelectRecordsAtPrint", None)
    except Exception: prev_sel = None
    try: prev_rr = getattr(fmt, "RecordRange", None)
    except Exception: prev_rr = None
    try: prev_copies = fmt.PrintSetup.IdenticalCopiesOfLabel
    except Exception: prev_copies = None

    try:
        # Отключение БД и копии=1
        try: setattr(fmt, "UseDatabase", False)
        except Exception: pass
        try: setattr(fmt, "SelectRecordsAtPrint", False)
        except Exception: pass
        try: setattr(fmt, "RecordRange", "1")
        except Exception: pass
        try: fmt.PrintSetup.IdenticalCopiesOfLabel = 1
        except Exception: pass

        try:
            self.logger.log("КАЛИБРОВКА: печатаю 6 тестовых этикеток 'X'…")
        except Exception:
            pass

        # Произвольные подстановки 'X' (на случай видимых SubStrings)
        for k in ("DM","ShortName","NAME","ShortGTIN","PART_NUM","NUM","PROD_DATE","EXP_DATE"):
            try:
                fmt.SetNamedSubStringValue(k, "X")
            except Exception:
                try:
                    subs = getattr(fmt, "SubStrings", None)
                    if subs:
                        subs(k).Value = "X"
                except Exception:
                    pass

        # 6 отпечатков
        for _ in range(6):
            try:
                fmt.PrintOut(False, True)
            except Exception:
                try:
                    fmt.PrintOut(1, False, False)
                except Exception:
                    pass

        try:
            self.logger.log("КАЛИБРОВКА: завершена.")
        except Exception:
            pass

    finally:
        # Восстановление состояния
        try:
            if prev_copies is not None:
                fmt.PrintSetup.IdenticalCopiesOfLabel = prev_copies
        except Exception:
            pass
        try:
            if prev_use_db is not None:
                setattr(fmt, "UseDatabase", prev_use_db)
        except Exception:
            pass
        try:
            if prev_sel is not None:
                setattr(fmt, "SelectRecordsAtPrint", prev_sel)
        except Exception:
            pass
        try:
            if prev_rr is not None:
                setattr(fmt, "RecordRange", prev_rr)
        except Exception:
            pass
# --- End helper ---



def _patch__log_db_connections(self, fmt):
    """
    Пишем в лог сколько БД-подключений видит COM через разные коллекции.
    Никаких перепривязок — только лог.
    """
    total = 0
    details = []
    # 1) fmt.DatabaseConnections (частая коллекция)
    try:
        conns = getattr(fmt, "DatabaseConnections", None)
        if conns is not None:
            cnt = int(conns.Count)
            total += cnt
            details.append(f"DatabaseConnections={cnt}")
    except Exception as e:
        details.append(f"DatabaseConnections: err={e}")

    # 2) fmt.DatabaseSetup.DatabaseConnections (иногда нужна именно она)
    try:
        ds = getattr(fmt, "DatabaseSetup", None)
        if ds is not None:
            conns2 = getattr(ds, "DatabaseConnections", None)
            if conns2 is not None:
                cnt2 = int(conns2.Count)
                total += cnt2
                details.append(f"DatabaseSetup.DatabaseConnections={cnt2}")
    except Exception as e:
        details.append(f"DatabaseSetup.DatabaseConnections: err={e}")

    # 3) fmt.Databases (устаревшая, но в старых BarTender бывает)
    try:
        dbs = getattr(fmt, "Databases", None)
        if dbs is not None:
            cnt3 = int(dbs.Count)
            total += cnt3
            details.append(f"Databases={cnt3}")
    except Exception as e:
        details.append(f"Databases: err={e}")

    # Чтобы не суммировать одно и то же (коллекции могут дублировать),
    # выводим детально и "разумный максимум"
    msg = " ; ".join(details) if details else "нет коллекций"
    try:
        self.logger.log(f"DB connections (raw): {msg}")
        self.logger.log(f"DB connections (summary): ~{total} (сумма по коллекциям, может содержать дубли)")
    except Exception:
        pass
    return total, msg

def _patch__write_tmp_batch_rows(self, rows, tmp_path):
    """
    Записывает tmp_batch.csv с колонками из self.REQ_COLS.
    rows — массив словарей (обогащённых self._enrich).
    """
    cols = list(getattr(self, "REQ_COLS", ["ShortName","ShortGTIN","EXP_DATE","PROD_DATE","PART_NUM","DM","NUM"]))
    _os.makedirs(_os.path.dirname(tmp_path), exist_ok=True)
    with open(tmp_path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, delimiter=",", quoting=csv.QUOTE_MINIMAL)
        w.writeheader()
        for r in rows:
            # только необходимые столбцы
            out = {c: (r.get(c, "") or "") for c in cols}
            w.writerow(out)
    try:
        self.logger.log(f"tmp_batch.csv записан: {tmp_path}")
    except Exception:
        pass

def _patch__collect_range_rows(self):
    """
    Забираем диапазон из self.csv_rows по index/limit, делаем self._enrich,
    возвращаем массив 'enriched' словарей.
    """
    # ensure CSV loaded
    if not getattr(self, "csv_rows", None):
        try:
            if not getattr(self, "csv_path", ""):
                return []
            from __main__ import load_kontur_raw as _load_kontur_raw  # original helper
            self.csv_rows = _load_kontur_raw(self.csv_path)
        except Exception as e:
            try:
                self.logger.err(f"CSV ошибка: {e}")
            except Exception:
                pass
            return []

    # pick range (1-based index)
    try:
        idx0 = max(0, int(self.index_entry.get().strip() or "1") - 1)
    except Exception:
        idx0 = 0

    limit_v = None
    try:
        t = (self.limit_entry.get() or "").strip()
        if t and t != "0":
            limit_v = max(1, int(t))
    except Exception:
        limit_v = None

    rows_all = self.csv_rows[idx0:] if not limit_v else self.csv_rows[idx0:idx0+limit_v]
    if not rows_all:
        return []

    enriched = []
    for i, base in enumerate(rows_all, start=1):
        idx1 = idx0 + i
        try:
            enr = self._enrich(base, idx1)
        except Exception as e:
            enr = None
        if not enr:
            try: self.logger.err(f"Строка {idx1}: данные не сформированы — пропуск")
            except Exception: pass
            continue
        enriched.append(enr)
    return enriched

def _patch__print_range_one_job_via_csv(self):
    """
    Печать диапазона ОДНИМ заданием, используя уже привязанный к BTW tmp_batch.csv.
    НИКАКОГО перепривязывания — только логируем, что видит BarTender.
    """
    self.cancel_requested = False

    # 1) Собрать и записать tmp_batch.csv
    rows = _patch__collect_range_rows(self)
    if not rows:
        from tkinter import messagebox as mb
        mb.showerror("Печать N шт", "Диапазон пуст — нечего печатать.")
        return

    # всегда перезаписываем tmp перед печатью
    tmp_path = os.path.join("C:\\auto_print", "tmp_batch.csv")
    try:
        _patch__write_tmp_batch_rows(self, rows, tmp_path)
    except Exception as e:
        try: self.logger.err(f"Не удалось записать tmp_batch: {e}")
        except Exception: pass
        return

    # 2) Определить формат и BTW (берём из первой строки)
    fmt_name = rows[0].get("_FORMAT", "16x16")
    btw = self._get_btw_for_format(fmt_name)
    if not btw:
        return

    # 3) Открыть BTW, залогировать NamedSubStrings и DB connections, НИЧЕГО НЕ ПЕРЕПРИВЯЗЫВАЯ
    try:
        fmt = self.bt.open_format(btw)
    except Exception as e:
        try: self.logger.err(f"Не удалось открыть BTW: {e}")
        except Exception: pass
        return

    # лог подполя
    try:
        names = [s.Name for s in fmt.NamedSubStrings]
        self.logger.log(f"NamedSubStrings: {names}")
    except Exception:
        pass

    # лог количества подключений БД
    try:
        total, msg = _patch__log_db_connections(self, fmt)
        if total <= 0:
            self.logger.err("В шаблоне не обнаружены DatabaseConnections. Убедись, что BTW привязан к C:\\auto_print\\tmp_batch.csv и сохранён.")
    except Exception:
        pass

    # 4) Запустить одно задание печати (через диалог или тихо) — БЕЗ перепривязки
    prn = self._get_printer()
    if not prn:
        try: fmt.Close(1)
        except Exception: pass
        return

    copies = self._get_copies()
    # назначим принтер и копии на всякий случай
    try: fmt.PrintSetup.Printer = prn
    except Exception: pass
    try: fmt.PrintSetup.PrinterName = prn
    except Exception: pass
    # === CALIB & PACK (one-job, тихий путь) ===
    do_cal = False
    for name in ("calib_var","calibrate_var","calibration_var","calib_check","calibrate_check"):
        v = getattr(self, name, None)
        if v is None: continue
        try:
            do_cal = bool(v.get()); break
        except Exception:
            try: do_cal = bool(v); break
            except Exception: pass

    if False:
        try:
            try:
                self.logger.log("КАЛИБРОВКА: пропускаю перед циклом (перенесена в каждый батч)")
            except Exception:
                pass
        except Exception as _e:
            try: self.logger.err(f"КАЛИБРОВКА: сбой — {_e}")
            except Exception: pass
    else:
        try: self.logger.log("КАЛИБРОВКА: перенесена в CSV — отдельная печать отключена.")
        except Exception: pass

    try:
        copies = self._get_copies()
    except Exception:
        copies = 1
    try:
        fmt.PrintSetup.IdenticalCopiesOfLabel = int(copies)
    except Exception:
        pass
    # === END CALIB & PACK ===

    try: fmt.PrintSetup.IdenticalCopiesOfLabel = int(copies)
    except Exception: pass

    # показывать диалог?
    try:
        prompt = bool(self.show_dialog_var.get())
    except Exception:
        prompt = False

    try:
        if False and prompt:
            self.logger.log("ИСПОЛЬЗУЮ ПЕЧАТЬ ПАКЕТАМИ (даже с диалогом)...")
            fmt.PrintOut(True, True)
        else:
            self.logger.log("[INFO] ONE-JOB: печать пакетами с подтверждением")
            import csv, time
            try:
                import tkinter as _tk, tkinter.messagebox as _mb
            except Exception:
                _tk = None; _mb = None

            master_csv = os.path.join(BASE_DIR, "tmp_batch.csv")

            def _get_pack_size():
                cands = ["batch_entry", "pack_n", "pack_by", "packet_by", "packet_size", "pack_var"]
                for name in cands:
                    try:
                        v = getattr(self, name)
                        val = int(v.get()) if hasattr(v, "get") else int(v)
                        if val > 0:
                            return val
                    except Exception:
                        pass
                return 0

            pack_size = _get_pack_size()

            with open(master_csv, "r", encoding="utf-8-sig", newline="") as rf:
                rdr = csv.reader(rf, delimiter=",", quotechar='"')
                rows_all = list(rdr)

            if not rows_all:
                self.logger.err("tmp_batch.csv пуст")
            else:
                header = rows_all[0]
                data_rows = rows_all[1:]
                total = len(data_rows)
                if pack_size <= 0 or pack_size >= total:
                    pack_size = total

                packs = (total + pack_size - 1) // pack_size
                for p in range(packs):
                    s = p * pack_size
                    e = min(total, s + pack_size)
                    chunk = data_rows[s:e]


                    # --- CSV-CAL: если галка включена и это первый батч — добавим 6 'X' строк в начало ---
                    def __x6_cal_enabled_any(self):
                        # Явно читаем известные имена чекбокса калибровки
                        names = ('calib_var','calibrate_var','calibration_var','calib_check','calibrate_check','calib_chk')
                        for _n in names:
                            try:
                                v = getattr(self, _n, None)
                                if v is None:
                                    continue
                                v = v.get() if hasattr(v, 'get') else v
                                if bool(v):
                                    return True
                            except Exception:
                                pass
                        return False

                    do_cal = __x6_cal_enabled_any(self)
                    add_x6 = bool(do_cal and len(chunk) > 0)

                    # Готовим калибровочную строку в формате текущего header
                    cal_rows = []
                    if add_x6:
                        # индекс поля _FORMAT, если есть
                        fmt_idx = None
                        try:
                            fmt_idx = header.index("_FORMAT")
                        except Exception:
                            fmt_idx = None
                        fmt0 = None
                        try:
                            if fmt_idx is not None and data_rows:
                                fmt0 = data_rows[0][fmt_idx] or "16x16"
                        except Exception:
                            fmt0 = "16x16"
                        # Собираем строку X под каждое имя колонки
                        def __x6_build_row():
                            rowx = []
                            for col in header:
                                if col.upper() == "NUM":
                                    rowx.append("1")
                                elif col == "ShortGTIN":
                                    rowx.append("000")
                                elif col == "_FORMAT":
                                    rowx.append(fmt0 or "16x16")
                                else:
                                    rowx.append("X")
                            return rowx
                        cal_rows = [__x6_build_row() for _ in range(6)]
                        try: self.logger.log("CSV-CAL: добавлю 6 'X' в файл батча (только для первого пакета)")
                        except Exception: pass
                    # --- END CSV-CAL ---
                    # Перезаписать CSV текущим батчем
                    with open(master_csv, "w", encoding="utf-8-sig", newline="") as wf:
                        w = csv.writer(wf, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)
                        w.writerow(header)
                        for row in cal_rows:
                            w.writerow(row)
                        for row in chunk:
                            w.writerow(row)
                    try:
                        self.logger.log(f"[PACK] tmp_batch.csv → строки {s+1}-{e} из {total}")
                    except Exception:
                        pass

                    # Перепривязать БД и включить печать из БД
                    try:
                        self._rangecsv_repoint_db(fmt, master_csv)
                    except Exception as _e:
                        try: self.logger.err(f"Rebind DB failed: {_e}")
                        except Exception: pass
                    try: fmt.UseDatabase = True
                    except Exception: pass
                    try: fmt.SelectRecordsAtPrint = False
                    except Exception: pass
                    try:
                        _do_cal = False
                        try:
                            names = ('calib_var','calibrate_var','calibration_var','calib_check','calibrate_check','calib_chk')
                            _do_cal = False
                            for _n in names:
                                try:
                                    _v = getattr(self, _n, None)
                                    if _v is None:
                                        continue
                                    _v = _v.get() if hasattr(_v, 'get') else _v
                                    if bool(_v):
                                        _do_cal = True
                                        break
                                except Exception:
                                    pass
                        except Exception:
                            pass
                        if _do_cal:
                            fmt.RecordRange = f"1-{len(chunk)+6}"
                        else:
                            fmt.RecordRange = f"1-{len(chunk)}"
                    except Exception: pass

                    # --- КАЛИБРОВКА: перед КАЖДЫМ батчем, если включена галочка ---

                    do_cal = False

                    for _name in ('calib_var','calibrate_var','calibration_var','calib_check','calibrate_check'):

                        _v = getattr(self, _name, None)

                        if _v is None: continue

                        try:

                            do_cal = bool(_v.get()); break

                        except Exception:

                            try: do_cal = bool(_v); break

                            except Exception: pass

                    if False:
                        try:

                            if self.logger: self.logger.log(f"КАЛИБРОВКА: батч {p+1}/{packs} — печатаю 6 тестовых этикеток 'X'…")

                        except Exception: pass

                        try:

                            _calibrate_fmt_safely(self, fmt)

                        except Exception as _e:

                            try:

                                if self.logger: self.logger.err(f"КАЛИБРОВКА: сбой — {_e}")

                            except Exception: pass


                    _show_dialog = False

                    try:

                        _show_dialog = bool(self.show_dialog_var.get())

                    except Exception:

                        try: _show_dialog = bool(self.default_show_dialog)

                        except Exception: _show_dialog = False

                    # Печать с ожиданием спулера (чтоб пакеты не слиплись)
                    try:
                        fmt.PrintOut(False, True) if self._dialog_flag() else fmt.PrintOut(False, False)
                    except TypeError:
                        fmt.PrintOut(int(copies), _show_dialog)
                        time.sleep(2)

                    # Подтверждение между батчами (кроме последнего)
                    # Учитываем «Одно задание»: подтверждать только если выключено
                    one_job = False
                    try:
                        one_job = bool(self.single_job_var.get())
                    except Exception:
                        try:
                            one_job = bool(self.single_job_var)
                        except Exception:
                            one_job = False
                    if p < packs - 1 and not one_job:
                        cont = True
                        try:
                            if _mb is not None:
                                cont = _mb.askyesno("Печать пакетами",
                                                    f"Батч {p+1}/{packs} ({s+1}-{e}, parent=self) напечатан.\nПродолжить следующий?")
                        except Exception:
                            pass
                        if not cont:
                            try: self.logger.log("[PACK] Пользователь остановил печать батчей.")
                            except Exception: pass
                            break
        self.logger.log("ONE-JOB OK")
    except Exception as e:
        self.logger.err(f"ONE-JOB печать провалилась: {e}\n{_tb.format_exc()}")
    finally:
        try: fmt.Close(1)
        except Exception: pass

def _patch__print_one_pdf_dialog(self):
    """
    'Печать в PDF' теперь сначала формирует актуальный tmp_batch.csv по текущему диапазону,
    затем печатает выбранную строку как раньше через диалог принтера PDF.
    """
    # 0) Всегда обновим tmp_batch по текущему диапазону — чтобы 30x20/16x16 были в свежем состоянии
    try:
        rows = _patch__collect_range_rows(self)
        if rows:
            tmp_path = os.path.join("C:\\auto_print", "tmp_batch.csv")
            _patch__write_tmp_batch_rows(self, rows, tmp_path)
    except Exception:
        pass

    # Вызываем оригинальную логику (переназначенная этим же патчем ниже)
    # Но мы полностью заменяем метод, поэтому просто повторяем надёжный путь:
    prn = self._get_printer()
    if not prn:
        from tkinter import messagebox as mb
        mb.showerror("Принтер", "Выбери принтер")
        return

    # взять индекс и копии
    try:
        idx = self._get_index()
        copies = self._get_copies()
    except Exception:
        from tkinter import messagebox as mb
        mb.showerror("Печать", "Некорректные индекс/копии")
        return

    # ensure csv rows
    if not getattr(self, "csv_rows", None):
        try:
            if getattr(self, "csv_path", ""):
                from __main__ import load_kontur_raw as _load_kontur_raw
                self.csv_rows = _load_kontur_raw(self.csv_path)
        except Exception as e:
            try: self.logger.err(f"CSV ошибка: {e}")
            except Exception: pass
            return
    if not self.csv_rows:
        from tkinter import messagebox as mb
        mb.showerror("CSV", "Нет данных")
        return

    base = self.csv_rows[max(1, min(idx, len(self.csv_rows))) - 1]
    enr = self._enrich(base, idx)
    if not enr:
        return

    fmt_name = enr.get("_FORMAT","16x16")
    path = self._get_btw_for_format(fmt_name)
    if not path:
        return

    fmt = self._prepare_btw(path, enr, prn)
    if not fmt:
        return

    # фиксируем принтер/копии
    for attr, val in (("PrinterName", prn), ("Printer", prn)):
        try: setattr(fmt.PrintSetup, attr, val)
        except Exception: pass
    for attr, val in (("IdenticalCopiesOfLabel", int(copies)),):
        try: setattr(fmt.PrintSetup, attr, val)
        except Exception: pass

    try:
        self.logger.log(f"Printer='{prn}', Copies={copies} → PDF dialog")
    except Exception:
        pass

    try:
        fmt.PrintOut(True, True)  # показать диалог, дождаться
        self.logger.log(f"Задание на PDF отправлено через диалог, копий: {copies}")
    except Exception as e:
        self.logger.err(f"PDF PrintOut error: {e}")
    finally:
        try: fmt.Close(1)
        except Exception: pass

# Attach patches
try:
    App._log_db_connections = _patch__log_db_connections
    App._print_range_one_job_via_csv = _patch__print_range_one_job_via_csv
    App._print_one_pdf_dialog = _patch__print_one_pdf_dialog
    # helper (not bound to class, just internal helpers called above)
except Exception as _e:
    # If App isn't defined yet for some reason, we'll ignore silently.
    pass
# === Monkey patches: DB log + tmp_batch regenerate, no-rebind (END) ===


if __name__ == "__main__":
    main()


# === SAFE CALIBRATION (append-only) ===
def _calibrate_fmt_safely(self, fmt):
    """Печатает 6 тех. этикеток 'X' без БД и копий."""
    try: prev_use_db = getattr(fmt, "UseDatabase", None)
    except Exception: prev_use_db = None
    try: prev_sel = getattr(fmt, "SelectRecordsAtPrint", None)
    except Exception: prev_sel = None
    try: prev_rr = getattr(fmt, "RecordRange", None)
    except Exception: prev_rr = None
    try: prev_copies = fmt.PrintSetup.IdenticalCopiesOfLabel
    except Exception: prev_copies = None
    try:
        try: setattr(fmt, "UseDatabase", False)
        except Exception: pass
        try: setattr(fmt, "SelectRecordsAtPrint", False)
        except Exception: pass
        try: setattr(fmt, "RecordRange", "1")
        except Exception: pass
        try: fmt.PrintSetup.IdenticalCopiesOfLabel = 1
        except Exception: pass
        for k in ("DM","ShortName","NAME","ShortGTIN","PART_NUM","NUM","PROD_DATE","EXP_DATE"):
            try:
                fmt.SetNamedSubStringValue(k, "X")
            except Exception:
                try:
                    subs = getattr(fmt, "SubStrings", None)
                    if subs: subs(k).Value = "X"
                except Exception: pass
        try: self.logger.log("КАЛИБРОВКА: печатаю 6 тестовых этикеток 'X'…")
        except Exception: pass
        for _ in range(6):
            try: fmt.PrintOut(False, True) if self._dialog_flag() else fmt.PrintOut(False, False)
            except Exception:
                try: fmt.PrintOut(1, False, False)
                except Exception: pass
        try: self.logger.log("КАЛИБРОВКА: завершена.")
        except Exception: pass
    finally:
        try:
            if prev_copies is not None:
                fmt.PrintSetup.IdenticalCopiesOfLabel = prev_copies
        except Exception: pass
        try:
            if prev_use_db is not None:
                setattr(fmt, "UseDatabase", prev_use_db)
        except Exception: pass
        try:
            if prev_sel is not None:
                setattr(fmt, "SelectRecordsAtPrint", prev_sel)
        except Exception: pass
        try:
            if prev_rr is not None:
                setattr(fmt, "RecordRange", prev_rr)
        except Exception: pass

try:
    App._calibrate_fmt_safely = _calibrate_fmt_safely
except Exception:
    pass
# === END SAFE CALIBRATION ===



def _pack_print_range_one_job_via_csv(self, fmt, csv_path, total_rows, copies):
    """
    Пакетная печать через CSV: режем tmp_batch.csv на куски по N строк
    и на каждый кусок вызываем PrintOut один раз.
    Это даёт 'пакет из N' и вопрос на продолжение между пакетами.
    """
    import csv as _csv, tempfile as _temp, shutil as _shutil

    pack_n = int(self._get_pack_size())
    try:
        self.logger.log(f"[DEBUG] total_rows={total_rows} pack_n={pack_n} copies={copies}")
    except Exception:
        pass
    if pack_n <= 0:
        # если пакет не задан – печать всего файла как есть
        try:
            fmt.PrintSetup.IdenticalCopiesOfLabel = int(copies)
        except Exception:
            pass
        (fmt.PrintOut(False, True) if self._dialog_flag() else fmt.PrintOut(False, False))
        return True

    # читаем все строки разом (небольшой csv)
    with open(csv_path, "r", encoding="utf-8") as f:
        all_lines = f.read().splitlines()
    if not all_lines:
        self.logger.log("[ERROR] CSV пустой — нечего печатать")
        return False

    header = all_lines[0]
    records = all_lines[1:]
    total = len(records)

    start_idx = 0
    pack_idx = 1
    while start_idx < total:
        end_idx = min(start_idx + pack_n, total)  # не включительно
        chunk = [header] + records[start_idx:end_idx]
        try:
            if bool(self.calib_var.get()):
                delim = ';' if (';' in header and header.count(';') >= header.count(',')) else ','
                cols = [c.strip() for c in header.split(delim)]
                dummy = [('1' if c.upper()=='NUM' else 'X') for c in cols]
                dummy_line = delim.join(dummy)
                cal_rows = [dummy_line]*6
                chunk = [header] + cal_rows + records[start_idx:end_idx]
        except Exception:
            pass

        # перезаписываем tmp_batch.csv этим куском
        with open(csv_path, "w", encoding="utf-8", newline="") as wf:
            wf.write("\n".join(chunk) + "\n")

        # лог
        try:
            self.logger.log(f"[PACK] {os.path.basename(csv_path)} → строки {start_idx+1}-{end_idx} из {total}")
        except Exception:
            pass

        # печать
        try:
            fmt.PrintSetup.IdenticalCopiesOfLabel = int(copies)
        except Exception:
            pass
        (fmt.PrintOut(False, True) if self._dialog_flag() else fmt.PrintOut(False, False))

        # подтверждение продолжения, если не конец
        if end_idx < total:
            if not self._confirm_continue_pack(pack_idx, start_idx+1, end_idx, total):
                break

        start_idx = end_idx
        pack_idx += 1

    # восстановим весь файл обратно (не обязательно, но аккуратно)
    with open(csv_path, "w", encoding="utf-8") as wf:
        wf.write("\n".join([header] + records) + "\n")

    return True



# ====== X6 CAL PATCH: safer writer overrides (minimal invasive) ======
def _x6__ap_write_tmp_batch_csv(self, tmp_csv_path, rows, req_cols=None):
    import csv, os
    req = req_cols or ["ShortName","ShortGTIN","EXP_DATE","PROD_DATE","PART_NUM","DM","NUM"]
    # prepend 6 X
    try:
        do_cal = bool(self.calib_var.get())
    except Exception:
        do_cal = False
    if do_cal and rows:
        def _is_cal(r, req_keys):
            try:
                for k in req_keys:
                    v = str(r.get(k, '') or '')
                    if k.upper() == 'NUM':
                        if v not in ('', '0', '1'):
                            return False
                    elif k == 'ShortGTIN':
                        if v not in ('', '0', '000'):
                            return False
                    else:
                        if v != 'X':
                            return False
                return True
            except Exception:
                return False
        first = rows[0] if len(rows) else None
        if first is None or not _is_cal(first, req):
            dummy = {k: ('1' if k.upper()=='NUM' else ('000' if k=='ShortGTIN' else 'X')) for k in req}
            try:
                fmt0 = rows[0].get('_FORMAT') or '16x16'
                dummy['_FORMAT'] = fmt0
            except Exception:
                pass
            rows = [dummy.copy() for _ in range(6)] + list(rows)
            try:
                self.logger.log("КАЛИБРОВКА: 6 строк 'X' добавлены в начало tmp_batch.csv (override)")
            except Exception:
                pass
    os.makedirs(os.path.dirname(tmp_csv_path), exist_ok=True)
    with open(tmp_csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=req, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)
        w.writeheader()
        for r in rows or []:
            w.writerow({k: (r.get(k, "") or "") for k in req})
    try:
        self.logger.log(f"tmp_batch.csv записан: {tmp_csv_path}")
    except Exception:
        pass
    return tmp_csv_path

def _x6__write_tmp_batch_csv(self, rows_enriched, path=os.path.join(BASE_DIR, "tmp_batch.csv")):
    import csv, os
    cols = getattr(self, "REQ_COLS", ["ShortName","ShortGTIN","EXP_DATE","PROD_DATE","PART_NUM","DM","NUM"])
    # prepend 6 X
    try:
        do_cal = bool(self.calib_var.get())
    except Exception:
        do_cal = False
    if do_cal and rows_enriched:
        def _is_cal2(r, req_keys):
            try:
                for k in req_keys:
                    v = str(r.get(k, '') or '')
                    if k.upper() == 'NUM':
                        if v not in ('', '0', '1'):
                            return False
                    elif k == 'ShortGTIN':
                        if v not in ('', '0', '000'):
                            return False
                    else:
                        if v != 'X':
                            return False
                return True
            except Exception:
                return False
        first = rows_enriched[0] if len(rows_enriched) else None
        if first is None or not _is_cal2(first, cols):
            dummy = {k: ('1' if k.upper()=='NUM' else ('000' if k=='ShortGTIN' else 'X')) for k in cols}
            try:
                fmt0 = rows_enriched[0].get('_FORMAT') or '16x16'
                dummy['_FORMAT'] = fmt0
            except Exception:
                pass
            rows_enriched = [dummy.copy() for _ in range(6)] + list(rows_enriched)
            try:
                self.logger.log("КАЛИБРОВКА: 6 строк 'X' добавлены в начало tmp_batch.csv (override rows_enriched)")
            except Exception:
                pass
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)
        w.writeheader()
        for enr in rows_enriched or []:
            w.writerow({k: (enr.get(k, "") or "") for k in cols})
    try:
        self.logger.log(f"tmp_batch.csv записан: {path} (строк={len(rows_enriched or [])})")
    except Exception:
        pass
    return path

# Bind overrides (monkey-patch) to App methods
try:
    App._ap_write_tmp_batch_csv = _x6__ap_write_tmp_batch_csv
except Exception:
    pass
try:
    App._write_tmp_batch_csv = _x6__write_tmp_batch_csv
except Exception:
    pass
# ====== END X6 CAL PATCH ======