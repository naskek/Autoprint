"""Data loading and transformation helpers."""
from __future__ import annotations

import csv
import datetime as dt
import os
import re
from calendar import monthrange
from typing import Dict, Iterable, List, Optional

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency on Windows
    load_workbook = None

try:
    import win32print
except Exception:  # pragma: no cover - optional dependency on Windows
    win32print = None


REQUIRED_COLUMNS = ["ShortName", "ShortGTIN", "EXP_DATE", "PROD_DATE", "PART_NUM", "DM", "NUM"]


def enum_printers() -> List[str]:
    if not win32print:
        return []
    flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
    try:
        return [name for (_f, _d, name, _c) in win32print.EnumPrinters(flags)]
    except Exception:
        return []


def only_digits(s: str | None) -> str:
    return re.sub(r"\D+", "", s or "")


def short_gtin(gtin: str) -> str:
    digits = only_digits(gtin)
    return digits[-3:].zfill(3) if digits else ""


def parse_date_ru(s: str) -> dt.date:
    """Strict DD.MM.YYYY parser."""
    s = (s or "").strip()
    if not s:
        raise ValueError("Пустая дата")
    try:
        return dt.datetime.strptime(s, "%d.%m.%Y").date()
    except Exception:
        raise ValueError(f"Неверный формат даты: '{s}', нужен ДД.ММ.ГГГГ")


def parse_prod_date(s: str) -> Optional[dt.date]:
    """Soft parser used for saved config values."""
    s = (s or "").strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def add_days(date_: dt.date, days: int) -> dt.date:
    return date_ + dt.timedelta(days=days)


def add_months(date_: dt.date, months: int) -> dt.date:
    y = date_.year + (date_.month - 1 + months) // 12
    m = (date_.month - 1 + months) % 12 + 1
    d = min(date_.day, monthrange(y, m)[1])
    return dt.date(y, m, d)


def add_years(date_: dt.date, years: int) -> dt.date:
    y = date_.year + years
    d = min(date_.day, monthrange(y, date_.month)[1])
    return dt.date(y, date_.month, d)


def load_kontur_raw(csv_path: str) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        rdr = csv.reader(f, delimiter="\t", quotechar='"')
        for parts in rdr:
            if not parts or all((p or "").strip() == "" for p in parts):
                continue
            parts = [(p or "").strip() for p in parts]
            while len(parts) < 3:
                parts.append("")
            dm, gtin, name = parts[0], parts[1], parts[2]
            if dm.strip().upper() == "DM" and (name.strip().upper() in ("NAME", "")):
                continue
            if not dm.strip():
                continue
            rows.append({"DM": dm, "GTIN": gtin, "NAME": name})
    return rows


def _norm(s: str) -> str:
    return re.sub(r"\s+", "", (s or "").strip().lower())


def _parse_shelf_life(text: str) -> Dict[str, Optional[int]]:
    """Parse shelf life description from Excel."""
    t = (text or "").strip().lower()
    res: Dict[str, Optional[int]] = {'days': None, 'months': None, 'years': None, 'weeks': None, 'raw': text}
    if not t:
        return res
    m = re.search(r"(\d+)", t)
    if not m:
        return res
    n = int(m.group(1))
    if any(w in t for w in ["год", "лет", "года"]):
        res['years'] = n
        return res
    if any(w in t for w in ["мес", "месяц", "месяцев", "месяца"]):
        res['months'] = n
        return res
    if any(w in t for w in ["нед", "неделя", "недели", "недель"]):
        res['weeks'] = n
        return res
    if any(w in t for w in ["сут", "дн"]):
        res['days'] = n
        return res
    res['months'] = n
    return res


def read_product_map(xlsx_path: str) -> Dict[str, Dict[str, object]]:
    """Load product map with GTIN metadata."""
    mapping: Dict[str, Dict[str, object]] = {}
    if not xlsx_path or not os.path.isfile(xlsx_path) or not load_workbook:
        return mapping
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        hdrs = [str(c.value or "").strip() for c in ws[1]]
        col = {h: i + 1 for i, h in enumerate(hdrs)}

        idx_gtin = col.get("GTIN") or col.get("ГТИН") or col.get("gtin")
        idx_pack = col.get("Упаковка") or col.get("УПАКОВКА")
        idx_shelf = col.get("Срок годности") or col.get("СРОК ГОДНОСТИ")
        idx_part = col.get("Шаблон партии") or col.get("ШАБЛОН ПАРТИИ") or col.get("Партия") or col.get("ПАРТИЯ")
        idx_short = (col.get("ShortName") or col.get("SHORTNAME") or col.get("Shortname") or col.get("shortname")
                     or col.get("Short Name") or col.get("SHORT NAME") or col.get("Короткое имя") or col.get("КОРОТКОЕ ИМЯ"))

        def cell(row, j: int | None) -> str:
            if not j:
                return ""
            v = row[j - 1].value
            return "" if v is None else str(v).strip()

        has_short = bool(idx_short)
        for row in ws.iter_rows(min_row=2):
            g = only_digits(cell(row, idx_gtin))
            if not g:
                continue
            pack = _norm(cell(row, idx_pack))
            shelf = cell(row, idx_shelf)
            fmt = ""
            if "ведро" in pack:
                fmt = "30x20"
            elif any(w in pack for w in ("банка", "туба")):
                fmt = "16x16"
            shelf_parsed = _parse_shelf_life(shelf)
            part_tpl = cell(row, idx_part)
            short_from_xlsx = cell(row, idx_short)
            mapping[g] = {"FORMAT": fmt, "SHELF": shelf_parsed, "PART_TEMPLATE": part_tpl, "SHORTNAME": short_from_xlsx}
        mapping["_HAS_SHORT_COL"] = has_short
        return mapping
    except Exception:
        return {}


def choose_format_for(gtin: str, product_map: Dict[str, Dict[str, object]], manual_choice: str) -> str:
    if manual_choice in ("16x16", "30x20"):
        return manual_choice
    info = product_map.get(only_digits(gtin), {})
    fmt = (info.get("FORMAT") or "") if info else ""
    return fmt if fmt in ("16x16", "30x20") else "16x16"


def make_part_num(prod_date: dt.date | None, part_template: str | None) -> str:
    yymmdd = prod_date.strftime("%y%m%d") if prod_date else ""
    if part_template:
        return part_template.replace("{DATE}", yymmdd)
    return yymmdd


def enrich_row(base_row: Dict[str, str], idx1: int, prod_date: dt.date, exp_days_override: Optional[int],
               product_map: Dict[str, Dict[str, object]], mode_choice: str) -> Dict[str, str]:
    dm, gtin, name = base_row.get("DM", ""), base_row.get("GTIN", ""), base_row.get("NAME", "")
    fmt = choose_format_for(gtin, product_map, mode_choice)
    info = product_map.get(only_digits(gtin), {})
    pd = prod_date or dt.date.today()

    exp_date = None
    shelf_info = info.get("SHELF") or {}
    log_shelf = ""
    if shelf_info:
        if shelf_info.get('years'):
            exp_date = add_years(pd, int(shelf_info['years']))
            log_shelf = f"years={shelf_info['years']}"
        elif shelf_info.get('months'):
            exp_date = add_months(pd, int(shelf_info['months']))
            log_shelf = f"months={shelf_info['months']}"
        elif shelf_info.get('weeks'):
            exp_date = add_days(pd, int(shelf_info['weeks']) * 7)
            log_shelf = f"weeks={shelf_info['weeks']}"
        elif shelf_info.get('days'):
            exp_date = add_days(pd, int(shelf_info['days']))
            log_shelf = f"days={shelf_info['days']}"
    if (exp_date is None) and (exp_days_override is not None):
        exp_date = add_days(pd, int(exp_days_override))
        log_shelf = f"override_days={exp_days_override}"

    part_num = make_part_num(pd, info.get("PART_TEMPLATE") or "")
    short_excel = (info.get("SHORTNAME") or "").strip()

    return {
        "DM": dm, "GTIN": gtin, "NAME": name,
        "ShortGTIN": short_gtin(gtin),
        "ShortName": short_excel if short_excel else name,
        "PROD_DATE": pd.strftime("%d.%m.%Y") if pd else "",
        "NUM": str(idx1),
        "EXP_DATE": exp_date.strftime("%d.%m.%Y") if exp_date else "",
        "PART_NUM": part_num,
        "_FORMAT": fmt,
        "_SHELF_LOG": log_shelf,
        "_SHORT_SRC": "Excel" if short_excel else "NAME"
    }
